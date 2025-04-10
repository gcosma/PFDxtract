import streamlit as st
import pyLDAvis
import pyLDAvis.sklearn
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import re
import requests
from bs4 import BeautifulSoup
import time
import urllib3
import io
import pdfplumber
import logging
import os
import zipfile
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import plotly.express as px
import plotly.graph_objects as go
from sklearn.feature_extraction.text import CountVectorizer, TfidfVectorizer
from sklearn.decomposition import LatentDirichletAllocation
from sklearn.preprocessing import normalize
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics import silhouette_score  # Added for semantic clustering
import networkx as nx
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from collections import Counter
from bs4 import BeautifulSoup, Tag
import json  # Added for JSON export functionality

# Initialize NLTK resources
import nltk

nltk.download("punkt")
nltk.download("stopwords")
nltk.download("averaged_perceptron_tagger")
import random
import string
import traceback
from datetime import datetime
from openpyxl.utils import get_column_letter
from sklearn.base import BaseEstimator, TransformerMixin
import scipy.sparse as sp
from typing import Union
from sklearn.metrics import (
    silhouette_score,
    calinski_harabasz_score,
    davies_bouldin_score,
)
from transformers import AutoTokenizer, AutoModel
import torch
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import re
from collections import Counter
from tqdm import tqdm
from collections import defaultdict  

class ThemeAnalyzer:
    def __init__(self, model_name="emilyalsentzer/Bio_ClinicalBERT"):
        """Initialize the BERT-based theme analyzer with sentence highlighting capabilities"""
        # Initialize transformer model and tokenizer
        st.info("Loading BERT model and tokenizer... This may take a moment.")
        self.tokenizer = AutoTokenizer.from_pretrained(model_name)
        self.model = AutoModel.from_pretrained(model_name)

        # Configuration settings
        self.config = {
            "base_similarity_threshold": 0.65,
            "keyword_match_weight": 0.3,
            "semantic_similarity_weight": 0.7,
            "max_themes_per_framework": 5,
            "context_window_size": 200,
        }

        # Initialize frameworks with themes
        self.frameworks = {
            "I-SIRch": self._get_isirch_framework(),
            "House of Commons": self._get_house_of_commons_themes(),
            "Extended Analysis": self._get_extended_themes(),
        }

        # Color mapping for themes
        self.theme_color_map = {}
        self.theme_colors = [
            "#FFD580",  # Light orange
            "#FFECB3",  # Light amber
            "#E1F5FE",  # Light blue
            "#E8F5E9",  # Light green
            "#F3E5F5",  # Light purple
            "#FFF3E0",  # Light orange
            "#E0F7FA",  # Light cyan
            "#F1F8E9",  # Light lime
            "#FFF8E1",  # Light yellow
            "#E8EAF6",  # Light indigo
            "#FCE4EC",  # Light pink
            "#F5F5DC",  # Beige
            "#E6E6FA",  # Lavender
            "#FFFACD",  # Lemon chiffon
            "#D1E7DD",  # Mint
            "#F8D7DA",  # Light red
            "#D1ECF1",  # Teal light
            "#FFF3CD",  # Light yellow
            "#D6D8D9",  # Light gray
            "#CFF4FC",  # Info light
        ]

        # Pre-assign colors to frameworks
        self._preassign_framework_colors()

    def _preassign_framework_colors(self):
        """Preassign colors to each framework for consistent coloring"""
        # Create a dictionary to track colors used for each framework
        framework_colors = {}

        # Assign colors to each theme in each framework
        for framework, themes in self.frameworks.items():
            for i, theme in enumerate(themes):
                theme_key = f"{framework}_{theme['name']}"
                # Assign color from the theme_colors list, cycling if needed
                color_idx = i % len(self.theme_colors)
                self.theme_color_map[theme_key] = self.theme_colors[color_idx]

    def get_bert_embedding(self, text, max_length=512):
        """Generate BERT embedding for text"""
        if not isinstance(text, str) or not text.strip():
            return np.zeros(768)

        # Tokenize with truncation
        inputs = self.tokenizer(
            text,
            return_tensors="pt",
            truncation=True,
            max_length=max_length,
            padding=True,
        )

        # Get embeddings
        with torch.no_grad():
            outputs = self.model(**inputs)

        # Use CLS token for sentence representation
        return outputs.last_hidden_state[:, 0, :].squeeze().numpy()

    def _get_contextual_embedding(self, text, keyword, window_size=100):
        """Get embedding for text surrounding the keyword occurrence"""
        if not isinstance(text, str) or not text.strip() or keyword not in text.lower():
            return self.get_bert_embedding(keyword)

        text_lower = text.lower()
        position = text_lower.find(keyword.lower())

        # Get context window
        start = max(0, position - window_size)
        end = min(len(text), position + len(keyword) + window_size)

        # Get contextual text
        context = text[start:end]
        return self.get_bert_embedding(context)

    def _calculate_combined_score(
        self, semantic_similarity, keyword_count, text_length
    ):
        """Calculate combined score that balances semantic similarity and keyword presence"""
        # Normalize keyword count by text length
        normalized_keyword_density = min(1.0, keyword_count / (text_length / 1000))

        # Weighted combination
        keyword_component = (
            normalized_keyword_density * self.config["keyword_match_weight"]
        )
        semantic_component = (
            semantic_similarity * self.config["semantic_similarity_weight"]
        )

        return keyword_component + semantic_component

    def _find_sentence_positions(self, text, keywords):
        """Find sentences containing keywords and return their positions"""
        if not isinstance(text, str) or not text.strip():
            return []

        # Split text into sentences
        sentence_endings = r"(?<=[.!?])\s+(?=[A-Z])"
        sentences = re.split(sentence_endings, text)

        # Track character positions and matched sentences
        positions = []
        current_pos = 0

        for sentence in sentences:
            if not sentence.strip():
                current_pos += len(sentence)
                continue

            # Check if any keyword is in this sentence
            sentence_lower = sentence.lower()
            matched_keywords = []

            for keyword in keywords:
                if keyword and len(keyword) >= 3 and keyword.lower() in sentence_lower:
                    # Check if it's a whole word using word boundaries
                    keyword_lower = keyword.lower()
                    pattern = r"\b" + re.escape(keyword_lower) + r"\b"
                    if re.search(pattern, sentence_lower):
                        matched_keywords.append(keyword)

            # If sentence contains any keywords, add to positions
            if matched_keywords:
                start_pos = current_pos
                end_pos = current_pos + len(sentence)
                # Join all matched keywords
                keywords_str = ", ".join(matched_keywords)
                positions.append((start_pos, end_pos, keywords_str, sentence))

            # Move to next position
            current_pos += len(sentence)

            # Account for sentence ending characters and whitespace
            if current_pos < len(text) and text[current_pos - 1] in ".!?":
                # Check for any whitespace after sentence ending
                space_count = 0
                while (
                    current_pos + space_count < len(text)
                    and text[current_pos + space_count].isspace()
                ):
                    space_count += 1
                current_pos += space_count

        return sorted(positions)
######
    def create_highlighted_html(self, text, theme_highlights):
        """Create HTML with sentences highlighted by theme with improved color consistency"""
        if not text or not theme_highlights:
            return text
        
        # Convert highlights to a flat list of positions
        all_positions = []
        for theme_key, positions in theme_highlights.items():
            theme_color = self._get_theme_color(theme_key)
            for pos_info in positions:
                # position format: (start_pos, end_pos, keywords_str, sentence)
                all_positions.append((
                    pos_info[0],  # start position
                    pos_info[1],  # end position
                    theme_key,    # theme key
                    pos_info[2],  # keywords string
                    pos_info[3],  # original sentence
                    theme_color   # theme color
                ))
        
        # Sort positions by start position
        all_positions.sort()
        
        # Merge overlapping sentences using primary theme's color
        merged_positions = []
        if all_positions:
            current = all_positions[0]
            for i in range(1, len(all_positions)):
                if all_positions[i][0] <= current[1]:  # Overlap
                    # Create a meaningful theme name combination
                    combined_theme = current[2] + " + " + all_positions[i][2]
                    combined_keywords = current[3] + " + " + all_positions[i][3]
                    
                    # Use the first theme's color for consistency
                    combined_color = current[5]
                    
                    # Update current with merged information
                    current = (
                        current[0],                # Keep original start position
                        max(current[1], all_positions[i][1]),  # Take the later end position
                        combined_theme,            # Combined theme names
                        combined_keywords,         # Combined keywords
                        current[4],                # Keep original sentence
                        combined_color             # Use the first theme's color
                    )
                else:
                    merged_positions.append(current)
                    current = all_positions[i]
            merged_positions.append(current)
        
        # Create highlighted text
        result = []
        last_end = 0
        
        for start, end, theme_key, keywords, sentence, color in merged_positions:
            # Add text before this highlight
            if start > last_end:
                result.append(text[last_end:start])
            
            # Simple styling with solid background color
            style = f"background-color:{color}; border:1px solid #666; border-radius:2px; padding:1px 2px;"
            tooltip = f"Theme: {theme_key}\nKeywords: {keywords}"
            result.append(f'<span style="{style}" title="{tooltip}">{text[start:end]}</span>')
            
            last_end = end
        
        # Add remaining text
        if last_end < len(text):
            result.append(text[last_end:])
        
        # Create HTML structure
        html_content = """
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <title>Highlighted Document Analysis</title>
            <style>
                body {
                    font-family: Arial, sans-serif;
                    margin: 20px;
                    line-height: 1.6;
                }
                table {
                    width: 100%;
                    border-collapse: collapse;
                    margin-bottom: 20px;
                }
                th, td {
                    border: 1px solid #ddd;
                    padding: 8px;
                    text-align: left;
                }
                th {
                    background-color: #f2f2f2;
                    font-weight: bold;
                }
                .paragraph-container {
                    border: 1px solid #ddd;
                    padding: 15px;
                    margin-bottom: 20px;
                    background-color: #f9f9f9;
                }
            </style>
        </head>
        <body>
            <h2>Document Theme Analysis</h2>
            
            <div class="paragraph-container">
                <h3>Highlighted Text</h3>
                <p>
        """
        
        # Add the highlighted text
        html_content += ''.join(result)
        
        html_content += """
                </p>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Framework</th>
                        <th>Theme</th>
                        <th>Color</th>
                        <th>Matched Keywords</th>
                        <th>Extracted Sentence</th>
                    </tr>
                </thead>
                <tbody>
        """
    
        # Add rows to HTML with color-coded backgrounds
        for start, end, theme_key, keywords, sentence, color in merged_positions:
            # Split theme key into framework and theme
            framework, theme = theme_key.split('_', 1) if '_' in theme_key else (theme_key, theme_key)
            
            html_content += f"""
                    <tr>
                        <td>{framework}</td>
                        <td>{theme}</td>
                        <td style="background-color: {color};">{color}</td>
                        <td>{keywords}</td>
                        <td>{sentence}</td>
                    </tr>
            """
    
        # Close HTML structure
        html_content += """
                </tbody>
            </table>
        </body>
        </html>
        """
    
        return html_content
    
    def convert_html_to_pdf(self, html_content, output_filename=None):
        """
        Convert the generated HTML to a PDF file
        
        Args:
            html_content (str): HTML content to convert
            output_filename (str, optional): Filename for the PDF. 
                                             If None, generates a timestamped filename.
        
        Returns:
            str: Path to the generated PDF file
        """
        try:
            # Import weasyprint
            from weasyprint import HTML, CSS
            from datetime import datetime
            import os
            
            # Generate default filename if not provided
            if output_filename is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"theme_analysis_{timestamp}.pdf"
            
            # Ensure the filename ends with .pdf
            if not output_filename.lower().endswith('.pdf'):
                output_filename += '.pdf'
            
            # Additional CSS to ensure proper PDF rendering
            additional_css = """
            @page {
                size: A4;
                margin: 1cm;
            }
            body {
                font-family: Arial, sans-serif;
                font-size: 12pt;
                line-height: 1.6;
            }
            table {
                width: 100%;
                border-collapse: collapse;
                margin-bottom: 20px;
            }
            th, td {
                border: 1px solid #ddd;
                padding: 8px;
                text-align: left;
            }
            .paragraph-container {
                page-break-inside: avoid;
            }
            """
            
            # Create PDF
            HTML(string=html_content).write_pdf(
                output_filename, 
                stylesheets=[CSS(string=additional_css)]
            )
            
            return output_filename
        
        except ImportError:
            # Handle case where weasyprint is not installed
            st.error("WeasyPrint is not installed. Please install it using 'pip install weasyprint'")
            return None
        except Exception as e:
            # Handle other potential errors
            st.error(f"Error converting HTML to PDF: {str(e)}")
            return None
        
    def _create_integrated_html_for_pdf(self, results_df, highlighted_texts):
        """
        Create a single integrated HTML file with all highlighted records, themes, and framework information
        that can be easily converted to PDF
        """
        from collections import defaultdict
        from datetime import datetime
    
        # Map report IDs to their themes
        report_themes = defaultdict(list)
        
        # Ensure all themes have unique colors
        self._ensure_unique_theme_colors(results_df)
    
        # Build the report data with consistent colors
        for _, row in results_df.iterrows():
            if "Record ID" in row and "Theme" in row and "Framework" in row:
                record_id = row["Record ID"]
                framework = row["Framework"]
                theme = row["Theme"]
                confidence = row.get("Confidence", "")
                score = row.get("Combined Score", 0)
                matched_keywords = row.get("Matched Keywords", "")
    
                # Get theme color from our mapping
                theme_key = f"{framework}_{theme}"
                theme_color = self._get_theme_color(theme_key)
    
                report_themes[record_id].append({
                    "framework": framework,
                    "theme": theme,
                    "confidence": confidence,
                    "score": score,
                    "keywords": matched_keywords,
                    "color": theme_color,
                    "theme_key": theme_key
                })
    
        # Create HTML content with modern styling
        html_content = """
            <!DOCTYPE html>
            <html>
            <head>
                <title>BERT Theme Analysis Report</title>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1">
                <style>
                    body { 
                        font-family: Arial, sans-serif; 
                        line-height: 1.6; 
                        margin: 0;
                        padding: 20px;
                        color: #333;
                        background-color: #f9f9f9;
                    }
                    h1 { 
                        color: #2c3e50; 
                        border-bottom: 3px solid #3498db; 
                        padding-bottom: 10px; 
                        margin-top: 30px;
                        font-weight: 600;
                    }
                    h2 { 
                        color: #2c3e50; 
                        margin-top: 30px; 
                        border-bottom: 2px solid #bdc3c7; 
                        padding-bottom: 5px; 
                        font-weight: 600;
                    }
                    h3 {
                        color: #34495e;
                        font-weight: 600;
                        margin-top: 20px;
                    }
                    .record-container { 
                        margin-bottom: 40px; 
                        background-color: white;
                        border-radius: 8px;
                        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
                        padding: 20px;
                        page-break-after: always; 
                    }
                    .highlighted-text { 
                        margin: 15px 0; 
                        padding: 15px; 
                        border-radius: 4px;
                        border: 1px solid #ddd; 
                        background-color: #fff; 
                        line-height: 1.7;
                    }
                    .theme-info { margin: 15px 0; }
                    .theme-info table { 
                        border-collapse: collapse; 
                        width: 100%; 
                        margin-top: 15px;
                        border-radius: 4px;
                        overflow: hidden;
                    }
                    .theme-info th, .theme-info td { 
                        border: 1px solid #ddd; 
                        padding: 12px; 
                        text-align: left; 
                    }
                    .theme-info th { 
                        background-color: #3498db; 
                        color: white;
                        font-weight: 600;
                    }
                    .theme-info tr:nth-child(even) { background-color: #f9f9f9; }
                    .theme-info tr:hover { background-color: #f1f1f1; }
                    .high-confidence { background-color: #D5F5E3; }  /* Light green */
                    .medium-confidence { background-color: #FCF3CF; } /* Light yellow */
                    .low-confidence { background-color: #FADBD8; }   /* Light red */
                    .report-header {
                        background-color: #3498db;
                        color: white;
                        padding: 30px;
                        text-align: center;
                        border-radius: 8px;
                        margin-bottom: 30px;
                    }
                    .summary-card {
                        background-color: white;
                        border-radius: 8px;
                        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
                        padding: 20px;
                        margin-bottom: 30px;
                        display: flex;
                        flex-wrap: wrap;
                        justify-content: space-between;
                    }
                    .summary-box {
                        flex: 1;
                        min-width: 200px;
                        padding: 15px;
                        text-align: center;
                        border-right: 1px solid #eee;
                    }
                    .summary-box:last-child {
                        border-right: none;
                    }
                    .summary-number {
                        font-size: 36px;
                        font-weight: bold;
                        color: #3498db;
                        margin-bottom: 10px;
                    }
                    .summary-label {
                        font-size: 14px;
                        color: #7f8c8d;
                        text-transform: uppercase;
                    }
                    .theme-color-box {
                        display: inline-block;
                        width: 20px;
                        height: 20px;
                        margin-right: 5px;
                        vertical-align: middle;
                        border: 1px solid #999;
                    }
                    .legend-container {
                        background-color: white;
                        border-radius: 8px;
                        box-shadow: 0 3px 10px rgba(0,0,0,0.1);
                        padding: 15px;
                        margin-bottom: 20px;
                    }
                    .legend-title {
                        font-weight: bold;
                        margin-bottom: 10px;
                    }
                    .legend-item {
                        display: flex;
                        align-items: center;
                        margin-bottom: 5px;
                    }
                    @media print {
                        .record-container { page-break-after: always; }
                        body { background-color: white; }
                        .record-container, .summary-card { box-shadow: none; }
                    }
                    
                    /* Define theme-specific CSS classes for consistency */
        """
    
        # Add dynamic CSS classes for each theme
        theme_keys = set()
        for _, row in results_df.iterrows():
            if "Framework" in row and "Theme" in row:
                theme_keys.add(f"{row['Framework']}_{row['Theme']}")
        
        for theme_key in theme_keys:
            color = self._get_theme_color(theme_key)
            safe_class_name = "theme-" + theme_key.replace(" ", "-").replace("(", "").replace(")", "").replace(",", "").replace(".", "").lower()
            html_content += f"""
                    .{safe_class_name} {{
                        background-color: {color} !important;
                    }}
            """
    
        html_content += """
                </style>
            </head>
            <body>
                <div class="report-header">
                    <h1>BERT Theme Analysis Results</h1>
                    <p>Generated on """ + datetime.now().strftime("%d %B %Y, %H:%M") + """</p>
                </div>
                
                <div class="summary-card">
                    <div class="summary-box">
                        <div class="summary-number">""" + str(len(highlighted_texts)) + """</div>
                        <div class="summary-label">Documents Analyzed</div>
                    </div>
                    <div class="summary-box">
                        <div class="summary-number">""" + str(len(results_df)) + """</div>
                        <div class="summary-label">Theme Identifications</div>
                    </div>
                    <div class="summary-box">
                        <div class="summary-number">""" + str(len(results_df["Framework"].unique())) + """</div>
                        <div class="summary-label">Frameworks</div>
                    </div>
                </div>
                
                <!-- Add legend explaining gradients -->
                <div class="legend-container">
                    <div class="legend-title">Theme Color Guide</div>
                    <div>When text contains multiple themes, a gradient background is used to show all applicable themes. Check the tooltip for details.</div>
                </div>
            """
    
        # Add framework summary
        html_content += """
                <h2>Framework Summary</h2>
                <table class="theme-info">
                    <tr>
                        <th>Framework</th>
                        <th>Number of Themes</th>
                        <th>Number of Documents</th>
                    </tr>
            """
    
        for framework in results_df["Framework"].unique():
            framework_results = results_df[results_df["Framework"] == framework]
            num_themes = len(framework_results["Theme"].unique())
            num_docs = len(framework_results["Record ID"].unique())
    
            html_content += f"""
                    <tr>
                        <td>{framework}</td>
                        <td>{num_themes}</td>
                        <td>{num_docs}</td>
                    </tr>
                """
    
        html_content += """
                </table>
            """
    
        # Add each record with its themes and highlighted text
        html_content += "<h2>Document Analysis</h2>"
    
        for record_id, themes in report_themes.items():
            if record_id in highlighted_texts:
                record_title = next(
                    (row["Title"] for _, row in results_df.iterrows() if row.get("Record ID") == record_id),
                    f"Document {record_id}"
                )
    
                html_content += f"""
                    <div class="record-container">
                        <h2>Document: {record_title}</h2>
                        
                        <div class="theme-info">
                            <h3>Identified Themes</h3>
                            <table>
                                <tr>
                                    <th>Framework</th>
                                    <th>Theme</th>
                                    <th>Confidence</th>
                                    <th>Score</th>
                                    <th>Matched Keywords</th>
                                    <th>Color</th>
                                </tr>
                    """
    
                # Add theme rows with consistent styling and no gradients
                for theme_info in sorted(themes, key=lambda x: (x["framework"], -x.get("score", 0))):
                    theme_color = theme_info["color"]
                    
                    html_content += f"""
                                <tr style="background-color: {theme_color};">
                                    <td>{theme_info['framework']}</td>
                                    <td>{theme_info['theme']}</td>
                                    <td>{theme_info.get('confidence', '')}</td>
                                    <td>{round(theme_info.get('score', 0), 3)}</td>
                                    <td>{theme_info.get('keywords', '')}</td>
                                    <td><div class="theme-color-box" style="background-color:{theme_color};"></div></td>
                                </tr>
                        """
    
                html_content += """
                            </table>
                        </div>
                        
                        <div class="highlighted-text">
                            <h3>Text with Highlighted Keywords</h3>
                    """
    
                # Add highlighted text
                html_content += highlighted_texts[record_id]
    
                html_content += """
                        </div>
                    </div>
                    """
    
        html_content += """
            </body>
            </html>
            """
    
        return html_content
    
    def _ensure_unique_theme_colors(self, results_df):
        """Ensure all themes have unique colors by checking and reassigning if needed"""
        from collections import defaultdict
        
        # First collect all theme keys
        theme_keys = set()
        for _, row in results_df.iterrows():
            if "Framework" in row and "Theme" in row:
                theme_key = f"{row['Framework']}_{row['Theme']}"
                theme_keys.add(theme_key)
        
        # Assign colors for any missing themes
        for theme_key in theme_keys:
            if theme_key not in self.theme_color_map:
                self._assign_unique_theme_color(theme_key)
        
        # Check for duplicate colors and fix them
        color_to_themes = defaultdict(list)
        for theme_key in theme_keys:
            color = self.theme_color_map[theme_key]
            color_to_themes[color].append(theme_key)
        
        # Reassign colors for themes with duplicates
        for color, duplicate_themes in color_to_themes.items():
            if len(duplicate_themes) > 1:
                # Keep the first theme's color, reassign others
                for theme_key in duplicate_themes[1:]:
                    self._assign_unique_theme_color(theme_key)
    
    def _assign_unique_theme_color(self, theme_key):
        """Assign a unique color to a theme, ensuring no duplicates"""
        # Get currently used colors
        used_colors = set(self.theme_color_map.values())
        
        # Find an unused color from our palette
        for color in self.theme_colors:
            if color not in used_colors:
                self.theme_color_map[theme_key] = color
                return
        
        # If all colors are used, generate a new unique color
        import random
        
        def random_hex_color():
            """Generate a random pastel color that's visually distinct"""
            # Higher base value (200) ensures lighter/pastel colors
            r = random.randint(180, 240)
            g = random.randint(180, 240)
            b = random.randint(180, 240)
            return f"#{r:02x}{g:02x}{b:02x}"
        
        # Generate colors until we find one that's not too similar to existing ones
        while True:
            new_color = random_hex_color()
            if new_color not in used_colors:
                self.theme_color_map[theme_key] = new_color
                break

    
    def _create_gradient_css(self, colors):
        """Create a CSS gradient string from a list of colors
        
        For 2 colors: simple diagonal gradient
        For 3+ colors: striped gradient with equal divisions
        """
        if len(colors) == 2:
            # Simple diagonal gradient for 2 colors
            return f"linear-gradient(135deg, {colors[0]} 50%, {colors[1]} 50%)"
        else:
            # Create striped gradient for 3+ colors
            stops = []
            segment_size = 100.0 / len(colors)
            
            for i, color in enumerate(colors):
                start = i * segment_size
                end = (i + 1) * segment_size
                
                # Add color stop
                stops.append(f"{color} {start:.1f}%")
                stops.append(f"{color} {end:.1f}%")
            
            return f"linear-gradient(135deg, {', '.join(stops)})"




######



    

    
    def _get_theme_color(self, theme_key):
        """Get a consistent color for a specific theme"""
        # If this theme already has an assigned color, use it
        if theme_key in self.theme_color_map:
            return self.theme_color_map[theme_key]

        # Extract framework and theme from the theme_key (format: "framework_theme")
        parts = theme_key.split("_", 1)
        framework = parts[0] if len(parts) > 0 else "unknown"

        # Count existing colors for this framework
        framework_count = sum(
            1
            for existing_key in self.theme_color_map
            if existing_key.startswith(framework + "_")
        )

        # Assign the next available color from our palette
        color_idx = framework_count % len(self.theme_colors)
        assigned_color = self.theme_colors[color_idx]

        # Store the assignment for future consistency
        self.theme_color_map[theme_key] = assigned_color
        return assigned_color

    def analyze_document(self, text):
        """Analyze document text for themes and highlight sentences containing theme keywords"""
        if not isinstance(text, str) or not text.strip():
            return {}, {}

        # Get full document embedding
        document_embedding = self.get_bert_embedding(text)
        text_length = len(text.split())

        framework_themes = {}
        theme_highlights = {}

        for framework_name, framework_theme_list in self.frameworks.items():
            # Track keyword matches across the entire document
            all_keyword_matches = []

            # First pass: identify all keyword matches and their contexts
            theme_matches = []
            for theme in framework_theme_list:
                # Find all sentence positions containing any matching keywords
                sentence_positions = self._find_sentence_positions(
                    text, theme["keywords"]
                )

                # Extract keywords from sentence positions
                keyword_matches = []
                match_contexts = []

                for _, _, keywords_str, _ in sentence_positions:
                    for keyword in keywords_str.split(", "):
                        if keyword not in keyword_matches:
                            keyword_matches.append(keyword)

                            # Get contextual embeddings for each keyword occurrence
                            context_embedding = self._get_contextual_embedding(
                                text, keyword, self.config["context_window_size"]
                            )
                            match_contexts.append(context_embedding)

                # Calculate semantic similarity with theme description
                theme_description = theme["name"] + ": " + ", ".join(theme["keywords"])
                theme_embedding = self.get_bert_embedding(theme_description)
                theme_doc_similarity = cosine_similarity(
                    [document_embedding], [theme_embedding]
                )[0][0]

                # Calculate context similarities if available
                context_similarities = []
                if match_contexts:
                    for context_emb in match_contexts:
                        sim = cosine_similarity([context_emb], [theme_embedding])[0][0]
                        context_similarities.append(sim)

                # Use max context similarity if available, otherwise use document similarity
                max_context_similarity = (
                    max(context_similarities) if context_similarities else 0
                )
                semantic_similarity = max(theme_doc_similarity, max_context_similarity)

                # Calculate combined score
                combined_score = self._calculate_combined_score(
                    semantic_similarity, len(keyword_matches), text_length
                )

                if (
                    keyword_matches
                    and combined_score >= self.config["base_similarity_threshold"]
                ):
                    theme_matches.append(
                        {
                            "theme": theme["name"],
                            "semantic_similarity": round(semantic_similarity, 3),
                            "combined_score": round(combined_score, 3),
                            "matched_keywords": ", ".join(keyword_matches),
                            "keyword_count": len(keyword_matches),
                            "sentence_positions": sentence_positions,  # Store sentence positions for highlighting
                        }
                    )

                    all_keyword_matches.extend(keyword_matches)

            # Sort by combined score
            theme_matches.sort(key=lambda x: x["combined_score"], reverse=True)

            # Limit number of themes
            top_theme_matches = theme_matches[: self.config["max_themes_per_framework"]]

            # Store theme matches and their highlighting info
            if top_theme_matches:
                # Count keywords to identify potential overlaps
                keyword_counter = Counter(all_keyword_matches)

                # Filter out themes with high keyword overlap and lower scores
                final_themes = []
                used_keywords = set()

                for theme_match in top_theme_matches:
                    # Check if this theme adds unique keywords
                    theme_keywords = set(theme_match["matched_keywords"].split(", "))
                    unique_keywords = theme_keywords - used_keywords

                    # If theme adds unique keywords or has high score, include it
                    if unique_keywords or theme_match["combined_score"] > 0.75:
                        # Store the theme data
                        theme_match_data = {
                            "theme": theme_match["theme"],
                            "semantic_similarity": theme_match["semantic_similarity"],
                            "combined_score": theme_match["combined_score"],
                            "matched_keywords": theme_match["matched_keywords"],
                            "keyword_count": theme_match["keyword_count"],
                        }
                        final_themes.append(theme_match_data)

                        # Store the highlighting positions separately
                        theme_key = f"{framework_name}_{theme_match['theme']}"
                        theme_highlights[theme_key] = theme_match["sentence_positions"]

                        used_keywords.update(theme_keywords)

                framework_themes[framework_name] = final_themes
            else:
                framework_themes[framework_name] = []

        return framework_themes, theme_highlights

    def _get_isirch_framework(self):
        """I-SIRCh framework themes mapped exactly to the official framework structure"""
        return [
            {
                "name": "External - Policy factor",
                "keywords": ["policy factor", "policy", "factor"],
            },
            {
                "name": "External - Societal factor",
                "keywords": ["societal factor", "societal", "factor"],
            },
            {
                "name": "External - Economic factor",
                "keywords": ["economic factor", "economic", "factor"],
            },
            {"name": "External - COVID ✓", "keywords": ["covid ✓", "covid"]},
            {
                "name": "External - Geographical factor (e.g. Location of patient)",
                "keywords": [
                    "geographical factor",
                    "geographical",
                    "factor",
                    "location of patient",
                ],
            },
            {
                "name": "Internal - Physical layout and Environment",
                "keywords": [
                    "physical layout and environment",
                    "physical",
                    "layout",
                    "environment",
                ],
            },
            {
                "name": "Internal - Acuity (e.g., capacity of the maternity unit as a whole)",
                "keywords": ["acuity", "capacity of the maternity unit as a whole"],
            },
            {
                "name": "Internal - Availability (e.g., operating theatres)",
                "keywords": ["availability", "operating theatres"],
            },
            {
                "name": "Internal - Time of day (e.g., night working or day of the week)",
                "keywords": ["time of day", "time", "night working or day of the week"],
            },
            {
                "name": "Organisation - Team culture factor (e.g., patient safety culture)",
                "keywords": [
                    "team culture factor",
                    "team",
                    "culture",
                    "factor",
                    "patient safety culture",
                ],
            },
            {
                "name": "Organisation - Incentive factor (e.g., performance evaluation)",
                "keywords": [
                    "incentive factor",
                    "incentive",
                    "factor",
                    "performance evaluation",
                ],
            },
            {"name": "Organisation - Teamworking", "keywords": ["teamworking"]},
            {
                "name": "Organisation - Communication factor",
                "keywords": ["communication factor", "communication", "factor"],
            },
            {
                "name": "Organisation - Communication factor - Between staff",
                "keywords": ["between staff", "between", "staff"],
            },
            {
                "name": "Organisation - Communication factor - Between staff and patient (verbal)",
                "keywords": [
                    "between staff and patient",
                    "between",
                    "staff",
                    "patient",
                    "verbal",
                ],
            },
            {"name": "Organisation - Documentation", "keywords": ["documentation"]},
            {
                "name": "Organisation - Escalation/referral factor (including fresh eyes reviews)",
                "keywords": [
                    "escalation/referral factor",
                    "escalation/referral",
                    "factor",
                    "including fresh eyes reviews",
                    "specialist referral",
                    "delay in escalation",
                    "specialist review",
                    "senior input",
                    "interdisciplinary referral",
                    "escalation delay",
                    "consultant opinion",
                ],
            },
            {
                "name": "Organisation - National and/or local guidance",
                "keywords": [
                    "national and/or local guidance",
                    "national",
                    "and/or",
                    "local",
                    "guidance",
                    "national screening",
                    "screening program",
                    "standard implementation",
                    "standardized screening",
                    "protocol adherence",
                ],
            },
            {
                "name": "Organisation - Language barrier",
                "keywords": ["language barrier", "language", "barrier"],
            },
            {
                "name": "Jobs/Task - Assessment, investigation, testing, screening (e.g., holistic review)",
                "keywords": [
                    "assessment, investigation, testing, screening",
                    "assessment,",
                    "investigation,",
                    "testing,",
                    "screening",
                    "holistic review",
                    "specimen",
                    "sample",
                    "laboratory",
                    "test result",
                    "abnormal finding",
                    "test interpretation",
                ],
            },
            {
                "name": "Jobs/Task - Care planning",
                "keywords": ["care planning", "care", "planning"],
            },
            {
                "name": "Jobs/Task - Dispensing, administering",
                "keywords": [
                    "dispensing, administering",
                    "dispensing,",
                    "administering",
                ],
            },
            {"name": "Jobs/Task - Monitoring", "keywords": ["monitoring"]},
            {
                "name": "Jobs/Task - Risk assessment",
                "keywords": ["risk assessment", "risk", "assessment"],
            },
            {
                "name": "Jobs/Task - Situation awareness (e.g., loss of helicopter view)",
                "keywords": [
                    "situation awareness",
                    "situation",
                    "awareness",
                    "loss of helicopter view",
                ],
            },
            {
                "name": "Jobs/Task - Obstetric review",
                "keywords": ["obstetric review", "obstetric", "review"],
            },
            {"name": "Technologies - Issues", "keywords": ["issues"]},
            {
                "name": "Technologies - Interpretation (e.g., CTG)",
                "keywords": ["interpretation", "ctg"],
            },
            {
                "name": "Person - Patient (characteristics and performance)",
                "keywords": ["patient", "characteristics and performance"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics",
                "keywords": ["characteristics", "patient characteristics"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Physical characteristics",
                "keywords": ["physical characteristics", "physical", "characteristics"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Psychological characteristics (e.g., stress, mental health)",
                "keywords": [
                    "psychological characteristics",
                    "psychological",
                    "characteristics",
                    "stress",
                    "mental health",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Language competence (English)",
                "keywords": [
                    "language competence",
                    "language",
                    "competence",
                    "english",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Disability (e.g., hearing problems)",
                "keywords": ["disability", "hearing problems"],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Training and education (e.g., attendance at ante-natal classes)",
                "keywords": [
                    "training and education",
                    "training",
                    "education",
                    "attendance at ante-natal classes",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Characteristics - Record of attendance (e.g., failure to attend antenatal classes)",
                "keywords": [
                    "record of attendance",
                    "record",
                    "attendance",
                    "failure to attend antenatal classes",
                ],
            },
            {
                "name": "Person - Patient (characteristics and performance) - Performance",
                "keywords": ["performance", "patient performance"],
            },
            {
                "name": "Person - Staff (characteristics and performance)",
                "keywords": ["staff", "characteristics and performance"],
            },
            {
                "name": "Person - Staff (characteristics and performance) - Characteristics",
                "keywords": ["characteristics", "staff characteristics"],
            },
            {
                "name": "Person - Staff (characteristics and performance) - Performance",
                "keywords": ["performance", "staff performance"],
            },
        ]

    def _get_house_of_commons_themes(self):
        """House of Commons themes mapped exactly to the official document"""
        return [
            {
                "name": "Communication",
                "keywords": [
                    "communication",
                    "dismissed",
                    "listened",
                    "concerns not taken seriously",
                    "concerns",
                    "seriously",
                ],
            },
            {
                "name": "Fragmented care",
                "keywords": [
                    "fragmented care",
                    "fragmented",
                    "care",
                    "spread",
                    "poorly",
                    "communicating",
                    "providers",
                    "no clear coordination",
                    "clear",
                    "coordination",
                ],
            },
            {
                "name": "Guidance gaps",
                "keywords": [
                    "guidance gaps",
                    "guidance",
                    "gaps",
                    "information",
                    "needs",
                    "optimal",
                    "minority",
                ],
            },
            {
                "name": "Pre-existing conditions and comorbidities",
                "keywords": [
                    "pre-existing conditions and comorbidities",
                    "pre-existing",
                    "conditions",
                    "comorbidities",
                    "overrepresented",
                    "ethnic",
                    "minority",
                    "contribute",
                    "higher",
                    "mortality",
                ],
            },
            {
                "name": "Inadequate maternity care",
                "keywords": [
                    "inadequate maternity care",
                    "inadequate",
                    "maternity",
                    "care",
                    "individualized",
                    "culturally",
                    "sensitive",
                ],
            },
            {
                "name": "Care quality and access issues",
                "keywords": [
                    "microaggressions and racism",
                    "microaggressions",
                    "racism",
                    "implicit/explicit",
                    "impacts",
                    "access",
                    "treatment",
                    "quality",
                    "stereotyping",
                ],
            },
            {
                "name": "Socioeconomic factors and deprivation",
                "keywords": [
                    "socioeconomic factors and deprivation",
                    "socioeconomic",
                    "factors",
                    "deprivation",
                    "links to poor outcomes",
                    "links",
                    "outcomes",
                    "minority",
                    "overrepresented",
                    "deprived",
                    "areas",
                ],
            },
            {
                "name": "Biases and stereotyping",
                "keywords": [
                    "biases and stereotyping",
                    "biases",
                    "stereotyping",
                    "perpetuation",
                    "stereotypes",
                    "providers",
                ],
            },
            {
                "name": "Consent/agency",
                "keywords": [
                    "consent/agency",
                    "consent",
                    "agency",
                    "informed consent",
                    "agency over care decisions",
                    "informed",
                    "decisions",
                ],
            },
            {
                "name": "Dignity/respect",
                "keywords": [
                    "dignity/respect",
                    "dignity",
                    "respect",
                    "neglectful",
                    "lacking",
                    "discrimination faced",
                    "discrimination",
                    "faced",
                ],
            },
        ]

    def _get_extended_themes(self):
        """Extended Analysis themes with unique concepts not covered in I-SIRCh or House of Commons frameworks"""
        return [
            {
                "name": "Procedural and Process Failures",
                "keywords": [
                    "procedure failure",
                    "process breakdown",
                    "protocol breach",
                    "standard violation",
                    "workflow issue",
                    "operational failure",
                    "process gap",
                    "procedural deviation",
                    "system failure",
                    "process error",
                    "workflow disruption",
                    "task failure",
                ],
            },
            {
                "name": "Medication safety",
                "keywords": [
                    "medication safety",
                    "medication",
                    "safety",
                    "drug error",
                    "prescription",
                    "drug administration",
                    "medication error",
                    "adverse reaction",
                    "medication reconciliation",
                ],
            },
            {
                "name": "Resource allocation",
                "keywords": [
                    "resource allocation",
                    "resource",
                    "allocation",
                    "resource management",
                    "resource constraints",
                    "prioritisation",
                    "resource distribution",
                    "staffing levels",
                    "staff shortage",
                    "budget constraints",
                ],
            },
            {
                "name": "Facility and Equipment Issues",
                "keywords": [
                    "facility",
                    "equipment",
                    "maintenance",
                    "infrastructure",
                    "device failure",
                    "equipment malfunction",
                    "equipment availability",
                    "technical failure",
                    "equipment maintenance",
                    "facility limitations",
                ],
            },
            {
                "name": "Emergency preparedness",
                "keywords": [
                    "emergency preparedness",
                    "emergency protocol",
                    "emergency response",
                    "crisis management",
                    "contingency planning",
                    "disaster readiness",
                    "emergency training",
                    "rapid response",
                ],
            },
            {
                "name": "Staff Wellbeing and Burnout",
                "keywords": [
                    "burnout",
                    "staff wellbeing",
                    "resilience",
                    "psychological safety",
                    "stress management",
                    "compassion fatigue",
                    "work-life balance",
                    "staff support",
                    "mental health",
                    "emotional burden",
                ],
            },
            {
                "name": "Ethical considerations",
                "keywords": [
                    "ethical dilemma",
                    "ethical decision",
                    "moral distress",
                    "ethical conflict",
                    "value conflict",
                    "ethics committee",
                    "moral judgment",
                    "conscientious objection",
                    "ethical framework",
                ],
            },
            {
                "name": "Diagnostic process",
                "keywords": [
                    "diagnostic error",
                    "misdiagnosis",
                    "delayed diagnosis",
                    "diagnostic uncertainty",
                    "diagnostic reasoning",
                    "differential diagnosis",
                    "diagnostic testing",
                    "diagnostic accuracy",
                    "test interpretation",
                ],
            },
            {
                "name": "Post-Event Learning and Improvement",
                "keywords": [
                    "incident learning",
                    "corrective action",
                    "improvement plan",
                    "feedback loop",
                    "lessons learned",
                    "action tracking",
                    "improvement verification",
                    "learning culture",
                    "incident review",
                    "recommendation implementation",
                    "systemic improvement",
                    "organisational learning",
                ],
            },
            {
                "name": "Electronic Health Record Issues",
                "keywords": [
                    "electronic health record",
                    "ehr issue",
                    "alert fatigue",
                    "interface design",
                    "copy-paste error",
                    "dropdown selection",
                    "clinical decision support",
                    "digital documentation",
                    "system integration",
                    "information retrieval",
                    "data entry error",
                    "electronic alert",
                ],
            },
            {
                "name": "Time-Critical Interventions",
                "keywords": [
                    "time-critical",
                    "delayed intervention",
                    "response time",
                    "golden hour",
                    "deterioration recognition",
                    "rapid response",
                    "timely treatment",
                    "intervention delay",
                    "time sensitivity",
                    "critical timing",
                    "delayed recognition",
                    "prompt action",
                    "urgent intervention",
                    "emergency response",
                    "time-sensitive decision",
                    "immediate action",
                    "rapid assessment",
                ],
            },
            {
                "name": "Human Factors and Cognitive Aspects",
                "keywords": [
                    "cognitive bias",
                    "situational awareness",
                    "attention management",
                    "visual perception",
                    "cognitive overload",
                    "decision heuristic",
                    "tunnel vision",
                    "confirmation bias",
                    "fixation error",
                    "anchoring bias",
                    "memory limitation",
                    "cognitive fatigue",
                    "isolation decision-making",
                    "clinical confidence",
                    "professional authority",
                    "hierarchical barriers",
                    "professional autonomy",
                ],
            },
            {
                "name": "Service Design and Patient Flow",
                "keywords": [
                    "service design",
                    "patient flow",
                    "care pathway",
                    "bottleneck",
                    "patient journey",
                    "waiting time",
                    "system design",
                    "process mapping",
                    "patient transfer",
                    "capacity planning",
                    "workflow design",
                    "service bottleneck",
                ],
            },
            {
                "name": "Maternal and Neonatal Risk Factors",
                "keywords": [
                    "maternal risk",
                    "pregnancy complication",
                    "obstetric risk",
                    "neonatal risk",
                    "fetal risk",
                    "gestational diabetes",
                    "preeclampsia",
                    "placental issue",
                    "maternal age",
                    "parity",
                    "previous cesarean",
                    "multiple gestation",
                    "fetal growth restriction",
                    "prematurity",
                    "congenital anomaly",
                    "birth asphyxia",
                    "maternal obesity",
                    "maternal hypertension",
                    "maternal infection",
                    "obstetric hemorrhage",
                    "maternal cardiac",
                    "thromboembolism",
                ],
            },
            {
                "name": "Private vs. NHS Care Integration",
                "keywords": [
                    "private care",
                    "private midwife",
                    "private provider",
                    "NHS interface",
                    "care transition",
                    "private-public interface",
                    "independent provider",
                    "private consultation",
                    "private-NHS coordination",
                    "privately arranged care",
                    "independent midwife",
                    "cross-system communication",
                ],
            },
            {
                "name": "Peer Support and Supervision",
                "keywords": [
                    "peer support",
                    "collegial support",
                    "professional isolation",
                    "clinical supervision",
                    "peer review",
                    "case discussion",
                    "professional feedback",
                    "unsupported decision",
                    "lack of collegiality",
                    "professional network",
                    "mentoring",
                    "supervision",
                ],
            },
            {
                "name": "Diagnostic Testing and Specimen Handling",
                "keywords": [
                    "specimen",
                    "sample",
                    "test result",
                    "laboratory",
                    "analysis",
                    "interpretation",
                    "abnormal finding",
                    "discolored",
                    "contamination",
                    "collection",
                    "processing",
                    "transportation",
                    "storage",
                    "labeling",
                    "amniocentesis",
                    "blood sample",
                ],
            },
        ]

    # New methods to add
    def _get_confidence_label(self, score):
        """Convert numerical score to confidence label"""
        if score >= 0.7:
            return "High"
        elif score >= 0.5:
            return "Medium"
        else:
            return "Low"

    # First, we need to modify the theme analyzer's create_detailed_results method
    # to store the matched sentences with each theme detection

    def create_detailed_results(self, data, content_column="Content"):
        """
        Analyze multiple documents and create detailed results with progress tracking.

        Args:
            data (pd.DataFrame): DataFrame containing documents
            content_column (str): Name of the column containing text to analyze

        Returns:
            Tuple[pd.DataFrame, Dict]: (Results DataFrame, Dictionary of highlighted texts)
        """
        import streamlit as st

        results = []
        highlighted_texts = {}

        # Create progress tracking elements
        progress_bar = st.progress(0)
        status_text = st.empty()
        doc_count_text = st.empty()

        # Calculate total documents to process
        total_docs = len(data)
        doc_count_text.text(f"Processing 0/{total_docs} documents")

        # Process each document
        for idx, (i, row) in enumerate(data.iterrows()):
            # Update progress
            progress = (idx + 1) / total_docs
            progress_bar.progress(progress)
            status_text.text(
                f"Analyzing document {idx + 1}/{total_docs}: {row.get('Title', f'Document {i}')}"
            )

            # Skip empty content
            if pd.isna(row[content_column]) or row[content_column] == "":
                continue

            content = str(row[content_column])

            # Analyze themes and get highlights
            framework_themes, theme_highlights = self.analyze_document(content)

            # Create highlighted HTML for this document
            highlighted_html = self.create_highlighted_html(content, theme_highlights)
            highlighted_texts[i] = highlighted_html

            # Store results for each theme
            theme_count = 0
            for framework_name, themes in framework_themes.items():
                for theme in themes:
                    theme_count += 1

                    # Extract matched sentences for this theme
                    matched_sentences = []
                    theme_key = f"{framework_name}_{theme['theme']}"
                    if theme_key in theme_highlights:
                        for (
                            start_pos,
                            end_pos,
                            keywords_str,
                            sentence,
                        ) in theme_highlights[theme_key]:
                            matched_sentences.append(sentence)

                    # Join sentences if there are any
                    matched_text = (
                        "; ".join(matched_sentences) if matched_sentences else ""
                    )

                    results.append(
                        {
                            "Record ID": i,
                            "Title": row.get("Title", f"Document {i}"),
                            "Framework": framework_name,
                            "Theme": theme["theme"],
                            "Confidence": self._get_confidence_label(
                                theme["combined_score"]
                            ),
                            "Combined Score": theme["combined_score"],
                            "Semantic_Similarity": theme["semantic_similarity"],
                            "Matched Keywords": theme["matched_keywords"],
                            "Matched Sentences": matched_text,  # Add matched sentences to results
                        }
                    )

            # Update documents processed count with theme info
            doc_count_text.text(
                f"Processed {idx + 1}/{total_docs} documents. Found {theme_count} themes in current document."
            )

        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()

        # Final count update
        if results:
            doc_count_text.text(
                f"Completed analysis of {total_docs} documents. Found {len(results)} total themes."
            )
        else:
            doc_count_text.text(
                f"Completed analysis, but no themes were identified in the documents."
            )

        # Create results DataFrame
        results_df = pd.DataFrame(results) if results else pd.DataFrame()

        return results_df, highlighted_texts

    # Now let's modify the export_to_excel function to ensure it includes matched sentences

    def create_comprehensive_pdf(
        self, results_df, highlighted_texts, output_filename=None
    ):
        """
        Create a comprehensive PDF report with analysis results

        Args:
            results_df (pd.DataFrame): Results DataFrame
            highlighted_texts (Dict): Dictionary of highlighted texts
            output_filename (str, optional): Output filename

        Returns:
            str: Path to the created PDF file
        """
        from matplotlib.backends.backend_pdf import PdfPages
        from datetime import datetime
        import matplotlib.pyplot as plt
        import matplotlib.gridspec as gridspec
        import pandas as pd
        import numpy as np
        import os
        import tempfile
        from matplotlib.colors import LinearSegmentedColormap
        from matplotlib.patches import Patch

        # Generate default filename if not provided
        if output_filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"theme_analysis_report_{timestamp}.pdf"

        # Use a tempfile for matplotlib to avoid file conflicts
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmpfile:
            temp_pdf_path = tmpfile.name

        # Create PDF with matplotlib
        with PdfPages(temp_pdf_path) as pdf:
            # Title page
            fig = plt.figure(figsize=(12, 10))
            plt.text(
                0.5,
                0.6,
                "BERT Theme Analysis Report",
                fontsize=28,
                ha="center",
                va="center",
                weight="bold",
            )
            plt.text(
                0.5,
                0.5,
                f"Generated on {datetime.now().strftime('%d %B %Y, %H:%M')}",
                fontsize=16,
                ha="center",
                va="center",
            )

            # Add a decorative header bar
            plt.axhline(y=0.75, xmin=0.1, xmax=0.9, color="#3366CC", linewidth=3)
            plt.axhline(y=0.35, xmin=0.1, xmax=0.9, color="#3366CC", linewidth=3)

            # Add framework names
            frameworks = self.frameworks.keys()
            framework_text = "Frameworks analyzed: " + ", ".join(frameworks)
            plt.text(
                0.5,
                0.3,
                framework_text,
                fontsize=14,
                ha="center",
                va="center",
                style="italic",
            )

            plt.axis("off")
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

            # Summary statistics page
            if not results_df.empty:
                # Create a summary page with charts
                fig = plt.figure(figsize=(12, 10))
                gs = gridspec.GridSpec(3, 2, height_ratios=[1, 2, 2])

                # Header
                ax_header = plt.subplot(gs[0, :])
                ax_header.text(
                    0.5,
                    0.5,
                    "Analysis Summary",
                    fontsize=20,
                    ha="center",
                    va="center",
                    weight="bold",
                )
                ax_header.axis("off")

                # Document count and metrics
                ax_metrics = plt.subplot(gs[1, 0])
                doc_count = len(highlighted_texts)
                theme_count = len(results_df)
                frameworks_count = len(results_df["Framework"].unique())

                metrics_text = (
                    f"Total Documents Analyzed: {doc_count}\n"
                    f"Total Theme Predictions: {theme_count}\n"
                    f"Unique Frameworks: {frameworks_count}\n"
                )

                if "Confidence" in results_df.columns:
                    confidence_counts = results_df["Confidence"].value_counts()
                    metrics_text += "\nConfidence Levels:\n"
                    for conf, count in confidence_counts.items():
                        metrics_text += f"  {conf}: {count} themes\n"

                ax_metrics.text(
                    0.1, 0.9, metrics_text, fontsize=12, va="top", linespacing=2
                )
                ax_metrics.axis("off")

                # Framework distribution chart
                ax_framework = plt.subplot(gs[1, 1])
                if not results_df.empty:
                    framework_counts = results_df["Framework"].value_counts()
                    bars = ax_framework.bar(
                        framework_counts.index,
                        framework_counts.values,
                        color=["#3366CC", "#DC3912", "#FF9900"],
                    )
                    ax_framework.set_title(
                        "Theme Distribution by Framework", fontsize=14
                    )
                    ax_framework.set_ylabel("Number of Themes")

                    # Add value labels on bars
                    for bar in bars:
                        height = bar.get_height()
                        ax_framework.text(
                            bar.get_x() + bar.get_width() / 2.0,
                            height + 0.1,
                            f"{height:d}",
                            ha="center",
                            fontsize=10,
                        )

                    ax_framework.spines["top"].set_visible(False)
                    ax_framework.spines["right"].set_visible(False)
                    plt.setp(ax_framework.get_xticklabels(), rotation=30, ha="right")

                # Themes by confidence chart
                ax_confidence = plt.subplot(gs[2, :])
                if "Confidence" in results_df.columns and "Theme" in results_df.columns:
                    # Prepare data for stacked bar chart
                    theme_conf_data = pd.crosstab(
                        results_df["Theme"], results_df["Confidence"]
                    )

                    # Select top themes by total count
                    top_themes = (
                        theme_conf_data.sum(axis=1)
                        .sort_values(ascending=False)
                        .head(10)
                        .index
                    )
                    theme_conf_data = theme_conf_data.loc[top_themes]

                    # Plot stacked bar chart
                    confidence_colors = {
                        "High": "#4CAF50",
                        "Medium": "#FFC107",
                        "Low": "#F44336",
                    }

                    # Get confidence levels present in the data
                    confidence_levels = list(theme_conf_data.columns)
                    colors = [
                        confidence_colors.get(level, "#999999")
                        for level in confidence_levels
                    ]

                    theme_conf_data.plot(
                        kind="barh",
                        stacked=True,
                        ax=ax_confidence,
                        color=colors,
                        figsize=(10, 6),
                    )

                    ax_confidence.set_title(
                        "Top Themes by Confidence Level", fontsize=14
                    )
                    ax_confidence.set_xlabel("Number of Documents")
                    ax_confidence.set_ylabel("Theme")

                    # Create custom legend
                    patches = [
                        Patch(
                            color=confidence_colors.get(level, "#999999"), label=level
                        )
                        for level in confidence_levels
                    ]
                    ax_confidence.legend(
                        handles=patches, title="Confidence", loc="upper right"
                    )

                    ax_confidence.spines["top"].set_visible(False)
                    ax_confidence.spines["right"].set_visible(False)
                else:
                    ax_confidence.axis("off")
                    ax_confidence.text(
                        0.5,
                        0.5,
                        "Confidence data not available",
                        fontsize=14,
                        ha="center",
                        va="center",
                    )

                plt.tight_layout()
                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

            # Framework-specific pages
            for framework_name in self.frameworks.keys():
                # Filter results for this framework
                framework_results = results_df[
                    results_df["Framework"] == framework_name
                ]

                if not framework_results.empty:
                    # Create a new page for the framework
                    fig = plt.figure(figsize=(12, 10))

                    # Title
                    plt.suptitle(
                        f"{framework_name} Framework Analysis",
                        fontsize=20,
                        y=0.95,
                        weight="bold",
                    )

                    # Theme counts
                    theme_counts = framework_results["Theme"].value_counts().head(15)

                    if not theme_counts.empty:
                        plt.subplot(111)
                        bars = plt.barh(
                            theme_counts.index[::-1],
                            theme_counts.values[::-1],
                            color="#5975A4",
                            alpha=0.8,
                        )

                        # Add value labels
                        for i, bar in enumerate(bars):
                            width = bar.get_width()
                            plt.text(
                                width + 0.3,
                                bar.get_y() + bar.get_height() / 2,
                                f"{width:d}",
                                va="center",
                                fontsize=10,
                            )

                        plt.xlabel("Number of Documents")
                        plt.ylabel("Theme")
                        plt.title(f"Top Themes in {framework_name}", pad=20)
                        plt.grid(axis="x", linestyle="--", alpha=0.7)
                        plt.tight_layout(rect=[0, 0, 1, 0.95])  # Adjust for suptitle

                    pdf.savefig(fig, bbox_inches="tight")
                    plt.close(fig)

            # Sample highlighted documents (text descriptions only)
            if highlighted_texts:
                # Create document summaries page
                fig = plt.figure(figsize=(12, 10))
                plt.suptitle(
                    "Document Analysis Summaries", fontsize=20, y=0.95, weight="bold"
                )

                # We'll show summaries of a few documents
                max_docs_to_show = min(3, len(highlighted_texts))
                docs_to_show = list(highlighted_texts.keys())[:max_docs_to_show]

                # Get theme counts for each document
                doc_summaries = []
                for doc_id in docs_to_show:
                    doc_themes = results_df[results_df["Record ID"] == doc_id]
                    theme_count = len(doc_themes)
                    frameworks = doc_themes["Framework"].unique()
                    doc_summaries.append(
                        {
                            "doc_id": doc_id,
                            "theme_count": theme_count,
                            "frameworks": ", ".join(frameworks),
                        }
                    )

                # Format as a table-like display
                plt.axis("off")
                table_text = "Document Analysis Summaries:\n\n"
                for i, summary in enumerate(doc_summaries):
                    doc_id = summary["doc_id"]
                    table_text += f"Document {i+1} (ID: {doc_id}):\n"
                    table_text += f"  • Identified Themes: {summary['theme_count']}\n"
                    table_text += f"  • Frameworks: {summary['frameworks']}\n\n"

                plt.text(0.1, 0.8, table_text, fontsize=12, va="top", linespacing=1.5)

                # Also add a note about the full HTML version
                note_text = (
                    "Note: A detailed HTML report with highlighted text excerpts has been\n"
                    "created alongside this PDF. The HTML version provides interactive\n"
                    "highlighting of theme-related sentences in each document."
                )
                plt.text(
                    0.1,
                    0.3,
                    note_text,
                    fontsize=12,
                    va="top",
                    linespacing=1.5,
                    style="italic",
                    bbox=dict(
                        facecolor="#F0F0F0",
                        edgecolor="#CCCCCC",
                        boxstyle="round,pad=0.5",
                    ),
                )

                pdf.savefig(fig, bbox_inches="tight")
                plt.close(fig)

        # Copy the temp file to the desired output filename
        import shutil

        shutil.copy2(temp_pdf_path, output_filename)

        # Clean up temp file
        try:
            os.unlink(temp_pdf_path)
        except:
            pass

        return output_filename

    def _preassign_framework_colors(self):
        """Preassign colors to each framework for consistent coloring"""
        # Create a dictionary to track colors used for each framework
        framework_colors = {}

        # Assign colors to each theme in each framework
        for framework, themes in self.frameworks.items():
            for i, theme in enumerate(themes):
                theme_key = f"{framework}_{theme['name']}"
                # Assign color from the theme_colors list, cycling if needed
                color_idx = i % len(self.theme_colors)
                self.theme_color_map[theme_key] = self.theme_colors[color_idx]

    def export_to_excel(df: pd.DataFrame) -> bytes:
        """
        Export DataFrame to Excel bytes with proper formatting, including matched sentences
        """
        try:
            if df is None or len(df) == 0:
                raise ValueError("No data available to export")

            # Create clean copy for export
            df_export = df.copy()

            # Format dates to UK format
            if "date_of_report" in df_export.columns:
                df_export["date_of_report"] = df_export["date_of_report"].dt.strftime(
                    "%d/%m/%Y"
                )

            # Handle list columns (like categories)
            for col in df_export.columns:
                if df_export[col].dtype == "object":
                    df_export[col] = df_export[col].apply(
                        lambda x: ", ".join(x)
                        if isinstance(x, list)
                        else str(x)
                        if pd.notna(x)
                        else ""
                    )

            # Create output buffer
            output = io.BytesIO()

            # Write to Excel
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df_export.to_excel(writer, sheet_name="Reports", index=False)

                # Get the worksheet
                worksheet = writer.sheets["Reports"]

                # Auto-adjust column widths
                for idx, col in enumerate(df_export.columns, 1):
                    # Set larger width for Matched Sentences column
                    if col == "Matched Sentences":
                        worksheet.column_dimensions[get_column_letter(idx)].width = 80
                    else:
                        max_length = max(
                            df_export[col].astype(str).apply(len).max(),
                            len(str(col)),
                        )
                        adjusted_width = min(max_length + 2, 50)
                        column_letter = get_column_letter(idx)
                        worksheet.column_dimensions[
                            column_letter
                        ].width = adjusted_width

                # Add filters to header row
                worksheet.auto_filter.ref = worksheet.dimensions

                # Freeze the header row
                worksheet.freeze_panes = "A2"

                # Set wrap text for Matched Sentences column
                matched_sent_col = next(
                    (
                        idx
                        for idx, col in enumerate(df_export.columns, 1)
                        if col == "Matched Sentences"
                    ),
                    None,
                )
                if matched_sent_col:
                    col_letter = get_column_letter(matched_sent_col)
                    for row in range(2, len(df_export) + 2):
                        cell = worksheet[f"{col_letter}{row}"]
                        cell.alignment = cell.alignment.copy(wrapText=True)
                        # Set row height to accommodate wrapped text
                        worksheet.row_dimensions[row].height = 60

            # Get the bytes value
            output.seek(0)
            return output.getvalue()

        except Exception as e:
            logging.error(f"Error exporting to Excel: {e}", exc_info=True)
            raise Exception(f"Failed to export data to Excel: {str(e)}")

class BM25Vectorizer(BaseEstimator, TransformerMixin):
    """BM25 vectorizer implementation"""

    def __init__(
        self,
        k1: float = 1.5,
        b: float = 0.75,
        max_features: Optional[int] = None,
        min_df: Union[int, float] = 1,
        max_df: Union[int, float] = 1.0,
    ):
        self.k1 = k1
        self.b = b
        self.max_features = max_features
        self.min_df = min_df
        self.max_df = max_df

        self.count_vectorizer = CountVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            stop_words="english",
        )

    def fit(self, raw_documents: List[str], y=None):
        X = self.count_vectorizer.fit_transform(raw_documents)

        # Calculate document lengths
        self.doc_lengths = np.array(X.sum(axis=1)).flatten()
        self.avg_doc_length = np.mean(self.doc_lengths)

        # Calculate IDF scores
        n_samples = X.shape[0]
        df = np.bincount(X.indices, minlength=X.shape[1])
        df = np.maximum(df, 1)
        self.idf = np.log((n_samples - df + 0.5) / (df + 0.5) + 1.0)

        return self

    def transform(self, raw_documents: List[str]) -> sp.csr_matrix:
        X = self.count_vectorizer.transform(raw_documents)
        doc_lengths = np.array(X.sum(axis=1)).flatten()

        X = sp.csr_matrix(X)

        # Calculate BM25 scores
        for i in range(X.shape[0]):
            start_idx = X.indptr[i]
            end_idx = X.indptr[i + 1]

            freqs = X.data[start_idx:end_idx]
            length_norm = 1 - self.b + self.b * doc_lengths[i] / self.avg_doc_length

            # BM25 formula
            X.data[start_idx:end_idx] = (
                ((self.k1 + 1) * freqs) / (self.k1 * length_norm + freqs)
            ) * self.idf[X.indices[start_idx:end_idx]]

        return X

    def get_feature_names_out(self):
        return self.count_vectorizer.get_feature_names_out()


class WeightedTfidfVectorizer(BaseEstimator, TransformerMixin):
    """TF-IDF vectorizer with configurable weighting schemes"""

    def __init__(
        self,
        tf_scheme: str = "raw",
        idf_scheme: str = "smooth",
        norm: Optional[str] = "l2",
        max_features: Optional[int] = None,
        min_df: Union[int, float] = 1,
        max_df: Union[int, float] = 1.0,
    ):
        self.tf_scheme = tf_scheme
        self.idf_scheme = idf_scheme
        self.norm = norm
        self.max_features = max_features
        self.min_df = min_df
        self.max_df = max_df

        self.count_vectorizer = CountVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            stop_words="english",
        )

    def _compute_tf(self, X: sp.csr_matrix) -> sp.csr_matrix:
        if self.tf_scheme == "raw":
            return X
        elif self.tf_scheme == "log":
            X.data = np.log1p(X.data)
        elif self.tf_scheme == "binary":
            X.data = np.ones_like(X.data)
        elif self.tf_scheme == "augmented":
            max_tf = X.max(axis=1).toarray().flatten()
            max_tf[max_tf == 0] = 1
            for i in range(X.shape[0]):
                start = X.indptr[i]
                end = X.indptr[i + 1]
                X.data[start:end] = 0.5 + 0.5 * (X.data[start:end] / max_tf[i])
        return X

    def _compute_idf(self, X: sp.csr_matrix) -> np.ndarray:
        n_samples = X.shape[0]
        df = np.bincount(X.indices, minlength=X.shape[1])
        df = np.maximum(df, 1)

        if self.idf_scheme == "smooth":
            return np.log((n_samples + 1) / (df + 1)) + 1
        elif self.idf_scheme == "standard":
            return np.log(n_samples / df) + 1
        elif self.idf_scheme == "probabilistic":
            return np.log((n_samples - df) / df)

    def fit(self, raw_documents: List[str], y=None):
        X = self.count_vectorizer.fit_transform(raw_documents)
        self.idf_ = self._compute_idf(X)
        return self

    def transform(self, raw_documents: List[str]) -> sp.csr_matrix:
        X = self.count_vectorizer.transform(raw_documents)
        X = self._compute_tf(X)
        X = X.multiply(self.idf_)

        if self.norm:
            X = normalize(X, norm=self.norm, copy=False)

        return X

    def get_feature_names_out(self):
        return self.count_vectorizer.get_feature_names_out()


def get_vectorizer(
    vectorizer_type: str, max_features: int, min_df: float, max_df: float, **kwargs
) -> Union[TfidfVectorizer, BM25Vectorizer, WeightedTfidfVectorizer]:
    """Create and configure the specified vectorizer type"""

    if vectorizer_type == "tfidf":
        return TfidfVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            stop_words="english",
        )
    elif vectorizer_type == "bm25":
        return BM25Vectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            k1=kwargs.get("k1", 1.5),
            b=kwargs.get("b", 0.75),
        )
    elif vectorizer_type == "weighted":
        return WeightedTfidfVectorizer(
            max_features=max_features,
            min_df=min_df,
            max_df=max_df,
            tf_scheme=kwargs.get("tf_scheme", "raw"),
            idf_scheme=kwargs.get("idf_scheme", "smooth"),
        )
    else:
        raise ValueError(f"Unknown vectorizer type: {vectorizer_type}")


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s: %(message)s",
    handlers=[logging.FileHandler("app.log"), logging.StreamHandler()],
)

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Global headers for all requests
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.5",
    "Connection": "keep-alive",
    "Referer": "https://judiciary.uk/",
}

# Core utility functions
def make_request(
    url: str, retries: int = 3, delay: int = 2
) -> Optional[requests.Response]:
    """Make HTTP request with retries and delay"""
    for attempt in range(retries):
        try:
            time.sleep(delay)
            response = requests.get(url, headers=HEADERS, verify=False, timeout=30)
            response.raise_for_status()
            return response
        except Exception as e:
            if attempt == retries - 1:
                st.error(f"Request failed: {str(e)}")
                raise e
            time.sleep(delay * (attempt + 1))
    return None


def combine_document_text(row: pd.Series) -> str:
    """Combine all text content from a document"""
    text_parts = []

    # Add title and content
    if pd.notna(row.get("Title")):
        text_parts.append(str(row["Title"]))
    if pd.notna(row.get("Content")):
        text_parts.append(str(row["Content"]))

    # Add PDF contents
    pdf_columns = [
        col for col in row.index if col.startswith("PDF_") and col.endswith("_Content")
    ]
    for pdf_col in pdf_columns:
        if pd.notna(row.get(pdf_col)):
            text_parts.append(str(row[pdf_col]))

    return " ".join(text_parts)


def clean_text_for_modeling(text: str) -> str:
    """Clean text with enhanced noise removal"""
    if not isinstance(text, str):
        return ""

    try:
        # Convert to lowercase
        text = text.lower()

        # Remove URLs
        text = re.sub(r"http\S+|www\S+|https\S+", "", text)

        # Remove email addresses and phone numbers
        text = re.sub(r"\S+@\S+", "", text)
        text = re.sub(r"\b\d{3}[-.]?\d{3}[-.]?\d{4}\b", "", text)

        # Remove dates and times
        text = re.sub(
            r"\b\d{1,2}(?:st|nd|rd|th)?\s+(?:january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}\b",
            "",
            text,
            flags=re.IGNORECASE,
        )
        text = re.sub(r"\b\d{1,2}:\d{2}\b", "", text)
        text = re.sub(r"\b\d{1,2}/\d{1,2}/\d{2,4}\b", "", text)

        # Remove specific document-related terms
        text = re.sub(
            r"\b(?:ref|reference|case)(?:\s+no)?\.?\s*[-:\s]?\s*\w+[-\d]+\b",
            "",
            text,
            flags=re.IGNORECASE,
        )
        text = re.sub(
            r"\b(regulation|paragraph|section|subsection|article)\s+\d+\b",
            "",
            text,
            flags=re.IGNORECASE,
        )

        # Remove common legal document terms
        legal_terms = r"\b(coroner|inquest|hearing|evidence|witness|statement|report|dated|signed)\b"
        text = re.sub(legal_terms, "", text, flags=re.IGNORECASE)

        # Remove special characters and multiple spaces
        text = re.sub(r"[^a-z\s]", " ", text)
        text = re.sub(r"\s+", " ", text)

        # Remove very short words
        text = " ".join(word for word in text.split() if len(word) > 2)

        # Ensure minimum content length
        cleaned_text = text.strip()
        return cleaned_text if len(cleaned_text.split()) >= 3 else ""

    except Exception as e:
        logging.error(f"Error in text cleaning: {e}")
        return ""


def clean_text(text: str) -> str:
    """Clean text while preserving structure and metadata formatting"""
    if not text:
        return ""

    try:
        text = str(text)
        text = unicodedata.normalize("NFKD", text)

        replacements = {
            "â€™": "'",
            "â€œ": '"',
            "â€": '"',
            "â€¦": "...",
            'â€"': "-",
            "â€¢": "•",
            "Â": " ",
            "\u200b": "",
            "\uf0b7": "",
            "\u2019": "'",
            "\u201c": '"',
            "\u201d": '"',
            "\u2013": "-",
            "\u2022": "•",
        }

        for encoded, replacement in replacements.items():
            text = text.replace(encoded, replacement)

        text = re.sub(r"<[^>]+>", "", text)
        text = "".join(
            char if char.isprintable() or char == "\n" else " " for char in text
        )
        text = re.sub(r" +", " ", text)
        text = re.sub(r"\n+", "\n", text)

        return text.strip()

    except Exception as e:
        logging.error(f"Error in clean_text: {e}")
        return ""


def extract_metadata(content: str) -> dict:
    """
    Extract structured metadata from report content with improved category handling.

    Args:
        content (str): Raw report content

    Returns:
        dict: Extracted metadata including date, reference, names, categories, etc.
    """
    metadata = {
        "date_of_report": None,
        "ref": None,
        "deceased_name": None,
        "coroner_name": None,
        "coroner_area": None,
        "categories": [],
    }

    if not content:
        return metadata

    try:
        # Extract date patterns
        date_patterns = [
            r"Date of report:?\s*(\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+\s+\d{4})",
            r"Date of report:?\s*(\d{1,2}/\d{1,2}/\d{4})",
            r"DATED this (\d{1,2}(?:st|nd|rd|th)?\s+day of [A-Za-z]+\s+\d{4})",
            r"Date:?\s*(\d{1,2}\s+[A-Za-z]+\s+\d{4})",
        ]

        for pattern in date_patterns:
            date_match = re.search(pattern, content, re.IGNORECASE)
            if date_match:
                date_str = date_match.group(1)
                try:
                    if "/" in date_str:
                        date_obj = datetime.strptime(date_str, "%d/%m/%Y")
                    else:
                        date_str = re.sub(r"(?<=\d)(st|nd|rd|th)", "", date_str)
                        date_str = re.sub(r"day of ", "", date_str)
                        try:
                            date_obj = datetime.strptime(date_str, "%d %B %Y")
                        except ValueError:
                            date_obj = datetime.strptime(date_str, "%d %b %Y")

                    metadata["date_of_report"] = date_obj.strftime("%d/%m/%Y")
                    break
                except ValueError as e:
                    logging.warning(f"Invalid date format found: {date_str} - {e}")

        # Extract reference number
        ref_match = re.search(r"Ref(?:erence)?:?\s*([-\d]+)", content)
        if ref_match:
            metadata["ref"] = ref_match.group(1).strip()

        # Extract deceased name
        name_match = re.search(r"Deceased name:?\s*([^\n]+)", content)
        if name_match:
            metadata["deceased_name"] = clean_text(name_match.group(1)).strip()

        # Extract coroner details
        coroner_match = re.search(r"Coroner(?:\'?s)? name:?\s*([^\n]+)", content)
        if coroner_match:
            metadata["coroner_name"] = clean_text(coroner_match.group(1)).strip()

        area_match = re.search(r"Coroner(?:\'?s)? Area:?\s*([^\n]+)", content)
        if area_match:
            metadata["coroner_area"] = clean_text(area_match.group(1)).strip()

        # Extract categories with enhanced handling
        cat_match = re.search(
            r"Category:?\s*(.+?)(?=This report is being sent to:|$)",
            content,
            re.IGNORECASE | re.DOTALL,
        )
        if cat_match:
            category_text = cat_match.group(1).strip()

            # Normalize all possible separators to pipe
            category_text = re.sub(r"\s*[,;]\s*", "|", category_text)
            category_text = re.sub(r"[•·⋅‣⁃▪▫–—-]\s*", "|", category_text)
            category_text = re.sub(r"\s{2,}", "|", category_text)
            category_text = re.sub(r"\n+", "|", category_text)

            # Split into individual categories
            categories = category_text.split("|")
            cleaned_categories = []

            # Get standard categories for matching
            standard_categories = {cat.lower(): cat for cat in get_pfd_categories()}

            for cat in categories:
                # Clean and normalize the category
                cleaned_cat = clean_text(cat).strip()
                cleaned_cat = re.sub(r"&nbsp;", "", cleaned_cat)
                cleaned_cat = re.sub(
                    r"\s*This report.*$", "", cleaned_cat, flags=re.IGNORECASE
                )
                cleaned_cat = re.sub(r"[|,;]", "", cleaned_cat)

                # Only process non-empty categories
                if cleaned_cat and not re.match(r"^[\s|,;]+$", cleaned_cat):
                    # Try to match with standard categories
                    cat_lower = cleaned_cat.lower()

                    # Check for exact match first
                    if cat_lower in standard_categories:
                        cleaned_cat = standard_categories[cat_lower]
                    else:
                        # Try partial matching
                        for std_lower, std_original in standard_categories.items():
                            if cat_lower in std_lower or std_lower in cat_lower:
                                cleaned_cat = std_original
                                break

                    cleaned_categories.append(cleaned_cat)

            # Remove duplicates while preserving order
            seen = set()
            metadata["categories"] = [
                x
                for x in cleaned_categories
                if not (x.lower() in seen or seen.add(x.lower()))
            ]

        return metadata

    except Exception as e:
        logging.error(f"Error extracting metadata: {e}")
        return metadata


def get_pfd_categories() -> List[str]:
    """Get all available PFD report categories"""
    return [
        "Accident at Work and Health and Safety related deaths",
        "Alcohol drug and medication related deaths",
        "Care Home Health related deaths",
        "Child Death from 2015",
        "Community health care and emergency services related deaths",
        "Emergency services related deaths 2019 onwards",
        "Hospital Death Clinical Procedures and medical management related deaths",
        "Mental Health related deaths",
        "Other related deaths",
        "Police related deaths",
        "Product related deaths",
        "Railway related deaths",
        "Road Highways Safety related deaths",
        "Service Personnel related deaths",
        "State Custody related deaths",
        "Suicide from 2015",
        "Wales prevention of future deaths reports 2019 onwards",
    ]


# PDF handling functions
def save_pdf(
    pdf_url: str, base_dir: str = "pdfs"
) -> Tuple[Optional[str], Optional[str]]:
    """Download and save PDF, return local path and filename"""
    try:
        os.makedirs(base_dir, exist_ok=True)

        response = make_request(pdf_url)
        if not response:
            return None, None

        filename = os.path.basename(pdf_url)
        filename = re.sub(r"[^\w\-_\. ]", "_", filename)
        local_path = os.path.join(base_dir, filename)

        with open(local_path, "wb") as f:
            f.write(response.content)

        return local_path, filename

    except Exception as e:
        logging.error(f"Error saving PDF {pdf_url}: {e}")
        return None, None


def extract_pdf_content(pdf_path: str, chunk_size: int = 10) -> str:
    """Extract text from PDF file with memory management"""
    try:
        filename = os.path.basename(pdf_path)
        text_chunks = []

        with pdfplumber.open(pdf_path) as pdf:
            for i in range(0, len(pdf.pages), chunk_size):
                chunk = pdf.pages[i : i + chunk_size]
                chunk_text = "\n\n".join([page.extract_text() or "" for page in chunk])
                text_chunks.append(chunk_text)

        full_content = f"PDF FILENAME: {filename}\n\n{''.join(text_chunks)}"
        return clean_text(full_content)

    except Exception as e:
        logging.error(f"Error extracting PDF text from {pdf_path}: {e}")
        return ""


def get_report_content(url: str) -> Optional[Dict]:
    """Get full content from report page with improved PDF and response handling"""
    try:
        logging.info(f"Fetching content from: {url}")
        response = make_request(url)
        if not response:
            return None

        soup = BeautifulSoup(response.text, "html.parser")
        content = soup.find("div", class_="flow") or soup.find(
            "article", class_="single__post"
        )

        if not content:
            logging.warning(f"No content found at {url}")
            return None

        # Extract main report content
        paragraphs = content.find_all(["p", "table"])
        webpage_text = "\n\n".join(
            p.get_text(strip=True, separator=" ") for p in paragraphs
        )

        pdf_contents = []
        pdf_paths = []
        pdf_names = []
        pdf_types = []  # Track if PDF is main report or response

        # Find all PDF links with improved classification
        pdf_links = soup.find_all("a", href=re.compile(r"\.pdf$"))

        for pdf_link in pdf_links:
            pdf_url = pdf_link["href"]
            pdf_text = pdf_link.get_text(strip=True).lower()

            # Determine PDF type
            is_response = any(
                word in pdf_text.lower() for word in ["response", "reply"]
            )
            pdf_type = "response" if is_response else "report"

            if not pdf_url.startswith(("http://", "https://")):
                pdf_url = (
                    f"https://www.judiciary.uk{pdf_url}"
                    if not pdf_url.startswith("/")
                    else f"https://www.judiciary.uk/{pdf_url}"
                )

            pdf_path, pdf_name = save_pdf(pdf_url)

            if pdf_path:
                pdf_content = extract_pdf_content(pdf_path)
                pdf_contents.append(pdf_content)
                pdf_paths.append(pdf_path)
                pdf_names.append(pdf_name)
                pdf_types.append(pdf_type)

        return {
            "content": clean_text(webpage_text),
            "pdf_contents": pdf_contents,
            "pdf_paths": pdf_paths,
            "pdf_names": pdf_names,
            "pdf_types": pdf_types,
        }

    except Exception as e:
        logging.error(f"Error getting report content: {e}")
        return None


def scrape_page(url: str) -> List[Dict]:
    """Scrape a single page with improved PDF handling"""
    reports = []
    try:
        response = make_request(url)
        if not response:
            return []

        soup = BeautifulSoup(response.text, "html.parser")
        results_list = soup.find("ul", class_="search__list")

        if not results_list:
            logging.warning(f"No results list found on page: {url}")
            return []

        cards = results_list.find_all("div", class_="card")

        for card in cards:
            try:
                title_elem = card.find("h3", class_="card__title")
                if not title_elem:
                    continue

                title_link = title_elem.find("a")
                if not title_link:
                    continue

                title = clean_text(title_link.text)
                card_url = title_link["href"]

                if not card_url.startswith(("http://", "https://")):
                    card_url = f"https://www.judiciary.uk{card_url}"

                logging.info(f"Processing report: {title}")
                content_data = get_report_content(card_url)

                if content_data:
                    report = {
                        "Title": title,
                        "URL": card_url,
                        "Content": content_data["content"],
                    }

                    # Add PDF details with type classification
                    for i, (name, content, path, pdf_type) in enumerate(
                        zip(
                            content_data["pdf_names"],
                            content_data["pdf_contents"],
                            content_data["pdf_paths"],
                            content_data["pdf_types"],
                        ),
                        1,
                    ):
                        report[f"PDF_{i}_Name"] = name
                        report[f"PDF_{i}_Content"] = content
                        report[f"PDF_{i}_Path"] = path
                        report[f"PDF_{i}_Type"] = pdf_type

                    reports.append(report)
                    logging.info(f"Successfully processed: {title}")

            except Exception as e:
                logging.error(f"Error processing card: {e}")
                continue

        return reports

    except Exception as e:
        logging.error(f"Error fetching page {url}: {e}")
        return []


def get_total_pages(url: str) -> Tuple[int, int]:
    """
    Get total number of pages and total results count

    Returns:
        Tuple[int, int]: (total_pages, total_results)
    """
    try:
        response = make_request(url)
        if not response:
            logging.error(f"No response from URL: {url}")
            return 0, 0

        soup = BeautifulSoup(response.text, "html.parser")

        # First check for total results count
        total_results = 0
        results_header = soup.find("div", class_="search__header")
        if results_header:
            results_text = results_header.get_text()
            match = re.search(r"found (\d+) results?", results_text, re.IGNORECASE)
            if match:
                total_results = int(match.group(1))
                total_pages = (total_results + 9) // 10  # 10 results per page
                return total_pages, total_results

        # If no results header, check pagination
        pagination = soup.find("nav", class_="navigation pagination")
        if pagination:
            page_numbers = pagination.find_all("a", class_="page-numbers")
            numbers = [
                int(p.text.strip()) for p in page_numbers if p.text.strip().isdigit()
            ]
            if numbers:
                return max(numbers), len(numbers) * 10  # Approximate result count

        # If no pagination but results exist
        results = soup.find("ul", class_="search__list")
        if results and results.find_all("div", class_="card"):
            cards = results.find_all("div", class_="card")
            return 1, len(cards)

        return 0, 0

    except Exception as e:
        logging.error(f"Error in get_total_pages: {str(e)}")
        return 0, 0


def process_scraped_data(df: pd.DataFrame) -> pd.DataFrame:
    """Process and clean scraped data with metadata extraction"""
    try:
        if df is None or len(df) == 0:
            return pd.DataFrame()

        # Create a copy
        df = df.copy()

        # Extract metadata from Content field if it exists
        if "Content" in df.columns:
            # Process each row
            processed_rows = []
            for _, row in df.iterrows():
                # Start with original row data
                processed_row = row.to_dict()

                # Extract metadata using existing function
                content = str(row.get("Content", ""))
                metadata = extract_metadata(content)

                # Update row with metadata
                processed_row.update(metadata)
                processed_rows.append(processed_row)

            # Create new DataFrame from processed rows
            result = pd.DataFrame(processed_rows)
        else:
            result = df.copy()

        # Convert date_of_report to datetime with UK format handling
        if "date_of_report" in result.columns:

            def parse_date(date_str):
                if pd.isna(date_str):
                    return pd.NaT

                date_str = str(date_str).strip()

                # If already in DD/MM/YYYY format
                if re.match(r"\d{1,2}/\d{1,2}/\d{4}", date_str):
                    return pd.to_datetime(date_str, format="%d/%m/%Y")

                # Remove ordinal indicators
                date_str = re.sub(r"(\d)(st|nd|rd|th)", r"\1", date_str)

                # Try different formats
                formats = ["%Y-%m-%d", "%d-%m-%Y", "%d %B %Y", "%d %b %Y"]

                for fmt in formats:
                    try:
                        return pd.to_datetime(date_str, format=fmt)
                    except ValueError:
                        continue

                try:
                    return pd.to_datetime(date_str)
                except:
                    return pd.NaT

            result["date_of_report"] = result["date_of_report"].apply(parse_date)

        return result

    except Exception as e:
        logging.error(f"Error in process_scraped_data: {e}")
        return df


def get_category_slug(category: str) -> str:
    """Generate proper category slug for the website's URL structure"""
    if not category:
        return None

    # Create a slug exactly matching the website's format
    slug = (
        category.lower()
        .replace(" ", "-")
        .replace("&", "and")
        .replace("--", "-")
        .strip("-")
    )

    logging.info(f"Generated category slug: {slug} from category: {category}")
    return slug


def scrape_pfd_reports(
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    order: str = "relevance",
    start_page: int = 1,
    end_page: Optional[int] = None,
) -> List[Dict]:
    """
    Scrape PFD reports with enhanced progress tracking and proper pagination
    """
    all_reports = []
    base_url = "https://www.judiciary.uk/"

    try:
        # Initialize progress tracking
        progress_bar = st.progress(0)
        status_text = st.empty()
        report_count_text = st.empty()

        # Validate and prepare category
        category_slug = None
        if category:
            category_slug = (
                category.lower()
                .replace(" ", "-")
                .replace("&", "and")
                .replace("--", "-")
                .strip("-")
            )
            logging.info(f"Using category: {category}, slug: {category_slug}")

        # Construct initial search URL
        base_search_url = construct_search_url(
            base_url=base_url,
            keyword=keyword,
            category=category,
            category_slug=category_slug,
        )

        st.info(f"Searching at: {base_search_url}")

        # Get total pages and results count
        total_pages, total_results = get_total_pages(base_search_url)

        if total_results == 0:
            st.warning("No results found matching your search criteria")
            return []

        st.info(f"Found {total_results} matching reports across {total_pages} pages")

        # Apply page range limits
        start_page = max(1, start_page)  # Ensure start_page is at least 1
        if end_page is None:
            end_page = total_pages
        else:
            end_page = min(
                end_page, total_pages
            )  # Ensure end_page doesn't exceed total_pages

        if start_page > end_page:
            st.warning(f"Invalid page range: {start_page} to {end_page}")
            return []

        st.info(f"Scraping pages {start_page} to {end_page}")

        # Process each page in the specified range
        for current_page in range(start_page, end_page + 1):
            try:
                # Check if scraping should be stopped
                if (
                    hasattr(st.session_state, "stop_scraping")
                    and st.session_state.stop_scraping
                ):
                    st.warning("Scraping stopped by user")
                    break

                # Update progress
                progress = (current_page - start_page) / (end_page - start_page + 1)
                progress_bar.progress(progress)
                status_text.text(
                    f"Processing page {current_page} of {end_page} (out of {total_pages} total pages)"
                )

                # Construct current page URL
                page_url = construct_search_url(
                    base_url=base_url,
                    keyword=keyword,
                    category=category,
                    category_slug=category_slug,
                    page=current_page,
                )

                # Scrape current page
                page_reports = scrape_page(page_url)

                if page_reports:
                    # Deduplicate based on title and URL
                    existing_reports = {(r["Title"], r["URL"]) for r in all_reports}
                    new_reports = [
                        r
                        for r in page_reports
                        if (r["Title"], r["URL"]) not in existing_reports
                    ]

                    all_reports.extend(new_reports)
                    report_count_text.text(
                        f"Retrieved {len(all_reports)} unique reports so far..."
                    )

                # Add delay between pages
                time.sleep(2)

            except Exception as e:
                logging.error(f"Error processing page {current_page}: {e}")
                st.warning(
                    f"Error on page {current_page}. Continuing with next page..."
                )
                continue

        # Sort results if specified
        if order != "relevance":
            all_reports = sort_reports(all_reports, order)

        # Clear progress indicators
        progress_bar.empty()
        status_text.empty()
        report_count_text.empty()

        if all_reports:
            st.success(f"Successfully scraped {len(all_reports)} unique reports")
        else:
            st.warning("No reports were successfully retrieved")

        return all_reports

    except Exception as e:
        logging.error(f"Error in scrape_pfd_reports: {e}")
        st.error(f"An error occurred while scraping reports: {e}")
        return []


def construct_search_url(
    base_url: str,
    keyword: Optional[str] = None,
    category: Optional[str] = None,
    category_slug: Optional[str] = None,
    page: Optional[int] = None,
) -> str:
    """Constructs proper search URL with pagination"""
    # Start with base search URL
    url = f"{base_url}?s=&post_type=pfd"

    # Add category filter
    if category and category_slug:
        url += f"&pfd_report_type={category_slug}"

    # Add keyword search
    if keyword:
        url = f"{base_url}?s={keyword}&post_type=pfd"
        if category and category_slug:
            url += f"&pfd_report_type={category_slug}"

    # Add pagination
    if page and page > 1:
        url += f"&paged={page}"  # Changed from &page= to &paged= for proper pagination

    return url


def render_scraping_tab():
    """Render the scraping tab with a clean 2x2 filter layout and page range selection"""
    st.header("Scrape PFD Reports")

    # Initialize default values if not in session state
    if "init_done" not in st.session_state:
        st.session_state.init_done = True
        st.session_state["search_keyword_default"] = "report"
        st.session_state["category_default"] = ""
        st.session_state["order_default"] = "relevance"
        st.session_state["start_page_default"] = 1
        st.session_state["end_page_default"] = None

    if "scraped_data" in st.session_state and st.session_state.scraped_data is not None:
        st.success(f"Found {len(st.session_state.scraped_data)} reports")

        st.subheader("Results")
        st.dataframe(
            st.session_state.scraped_data,
            column_config={
                "URL": st.column_config.LinkColumn("Report Link"),
                "date_of_report": st.column_config.DateColumn(
                    "Date of Report", format="DD/MM/YYYY"
                ),
                "categories": st.column_config.ListColumn("Categories"),
            },
            hide_index=True,
        )

        show_export_options(st.session_state.scraped_data, "scraped")

    # Create the search form with page range selection
    with st.form("scraping_form"):
        # Create two rows with two columns each
        row1_col1, row1_col2 = st.columns(2)
        row2_col1, row2_col2 = st.columns(2)
        row3_col1, row3_col2 = st.columns(2)

        # First row
        with row1_col1:
            search_keyword = st.text_input(
                "Search keywords:",
                value=st.session_state.get("search_keyword_default", "report"),
                key="search_keyword",
                help="Do not leave empty, use 'report' or another search term",
            )

        with row1_col2:
            category = st.selectbox(
                "PFD Report type:",
                [""] + get_pfd_categories(),
                index=0,
                key="category",
                format_func=lambda x: x if x else "Select a category",
            )

        # Second row
        with row2_col1:
            order = st.selectbox(
                "Sort by:",
                ["relevance", "desc", "asc"],
                index=0,
                key="order",
                format_func=lambda x: {
                    "relevance": "Relevance",
                    "desc": "Newest first",
                    "asc": "Oldest first",
                }[x],
            )

        with row2_col2:
            # Get total pages for the query (preview)
            if search_keyword or category:
                base_url = "https://www.judiciary.uk/"

                # Prepare category slug
                category_slug = None
                if category:
                    category_slug = (
                        category.lower()
                        .replace(" ", "-")
                        .replace("&", "and")
                        .replace("--", "-")
                        .strip("-")
                    )

                # Create preview URL
                preview_url = construct_search_url(
                    base_url=base_url,
                    keyword=search_keyword,
                    category=category,
                    category_slug=category_slug,
                )

                try:
                    with st.spinner("Checking total pages..."):
                        total_pages, total_results = get_total_pages(preview_url)
                        if total_pages > 0:
                            st.info(
                                f"This search has {total_pages} pages with {total_results} results"
                            )
                            st.session_state["total_pages_preview"] = total_pages
                        else:
                            st.warning("No results found for this search")
                            st.session_state["total_pages_preview"] = 0
                except Exception as e:
                    st.error(f"Error checking pages: {str(e)}")
                    st.session_state["total_pages_preview"] = 0
            else:
                st.session_state["total_pages_preview"] = 0

        # Third row for page range
        with row3_col1:
            start_page = st.number_input(
                "Start page:",
                min_value=1,
                value=st.session_state.get("start_page_default", 1),
                key="start_page",
                help="First page to scrape (minimum 1)",
            )

        with row3_col2:
            end_page = st.number_input(
                "End page:",
                min_value=0,
                value=st.session_state.get("end_page_default", 0),
                key="end_page",
                help="Last page to scrape (0 for all pages)",
            )

        # Action buttons in a row
        button_col1, button_col2 = st.columns(2)
        with button_col1:
            submitted = st.form_submit_button("Search Reports")
        with button_col2:
            stop_scraping = st.form_submit_button("Stop Scraping")

    # Handle stop scraping
    if stop_scraping:
        st.session_state.stop_scraping = True
        st.warning("Scraping will be stopped after the current page completes...")
        return

    if submitted:
        try:
            # Store search parameters in session state
            st.session_state.last_search_params = {
                "keyword": search_keyword,
                "category": category,
                "order": order,
                "start_page": start_page,
                "end_page": end_page,
            }

            # Initialize stop_scraping flag
            st.session_state.stop_scraping = False

            # Convert end_page=0 to None (all pages)
            end_page_val = None if end_page == 0 else end_page

            # Perform scraping
            reports = scrape_pfd_reports(
                keyword=search_keyword,
                category=category if category else None,
                order=order,
                start_page=start_page,
                end_page=end_page_val,
            )

            if reports:
                # Process the data
                df = pd.DataFrame(reports)
                df = process_scraped_data(df)

                # Store in session state
                st.session_state.scraped_data = df
                st.session_state.data_source = "scraped"
                st.session_state.current_data = df

                # Trigger a rerun to refresh the page
                st.rerun()
            else:
                st.warning("No reports found matching your search criteria")

        except Exception as e:
            st.error(f"An error occurred: {e}")
            logging.error(f"Scraping error: {e}")
            return False


def render_topic_summary_tab(data: pd.DataFrame) -> None:
    """Topic analysis with weighting schemes and essential controls"""
    st.header("Topic Analysis & Summaries")
    st.markdown(
        """
    This analysis identifies key themes and patterns in the report contents, automatically clustering similar documents
    and generating summaries for each thematic group.
    """
    )

    # Show previous results if available
    if "topic_model" in st.session_state and st.session_state.topic_model is not None:
        st.sidebar.success("Previous analysis results available")
        if st.sidebar.button("View Previous Results"):
            render_summary_tab(st.session_state.topic_model, data)
            return

    st.subheader("Analysis Settings")

    # Text Processing
    col1, col2 = st.columns(2)

    with col1:
        # Vectorization method
        vectorizer_type = st.selectbox(
            "Vectorization Method",
            options=["tfidf", "bm25", "weighted"],
            help="Choose how to convert text to numerical features",
        )

        # Weighting Schemes
        if vectorizer_type == "weighted":
            tf_scheme = st.selectbox(
                "Term Frequency Scheme",
                options=["raw", "log", "binary", "augmented"],
                help="How to count term occurrences",
            )
            idf_scheme = st.selectbox(
                "Document Frequency Scheme",
                options=["smooth", "standard", "probabilistic"],
                help="How to weight document frequencies",
            )
        elif vectorizer_type == "bm25":
            k1 = st.slider(
                "Term Saturation (k1)",
                min_value=0.5,
                max_value=3.0,
                value=1.5,
                help="Controls term frequency impact",
            )
            b = st.slider(
                "Length Normalization (b)",
                min_value=0.0,
                max_value=1.0,
                value=0.75,
                help="Document length impact",
            )

    with col2:
        # Clustering Parameters
        min_cluster_size = st.slider(
            "Minimum Group Size",
            min_value=2,
            max_value=10,
            value=3,
            help="Minimum documents per theme",
        )

        max_features = st.slider(
            "Maximum Features",
            min_value=1000,
            max_value=10000,
            value=5000,
            step=1000,
            help="Number of terms to consider",
        )

    # Date range selection
    st.subheader("Date Range")
    date_col1, date_col2 = st.columns(2)
    with date_col1:
        start_date = st.date_input(
            "From",
            value=data["date_of_report"].min().date(),
            min_value=data["date_of_report"].min().date(),
            max_value=data["date_of_report"].max().date(),
        )

    with date_col2:
        end_date = st.date_input(
            "To",
            value=data["date_of_report"].max().date(),
            min_value=data["date_of_report"].min().date(),
            max_value=data["date_of_report"].max().date(),
        )

    # Category selection
    all_categories = set()
    for cats in data["categories"].dropna():
        if isinstance(cats, list):
            all_categories.update(cats)

    categories = st.multiselect(
        "Filter by Categories (Optional)",
        options=sorted(all_categories),
        help="Select specific categories to analyze",
    )

    # Analysis button
    analyze_clicked = st.button(
        "🔍 Analyze Documents", type="primary", use_container_width=True
    )

    if analyze_clicked:
        try:
            progress_bar = st.progress(0)
            status_text = st.empty()

            # Initialize
            progress_bar.progress(0.2)
            status_text.text("Processing documents...")
            initialize_nltk()

            # Filter data
            filtered_df = data.copy()

            # Apply date filter
            filtered_df = filtered_df[
                (filtered_df["date_of_report"].dt.date >= start_date)
                & (filtered_df["date_of_report"].dt.date <= end_date)
            ]

            # Apply category filter
            if categories:
                filtered_df = filter_by_categories(filtered_df, categories)

            # Remove empty content
            filtered_df = filtered_df[
                filtered_df["Content"].notna()
                & (filtered_df["Content"].str.strip() != "")
            ]

            if len(filtered_df) < min_cluster_size:
                progress_bar.empty()
                status_text.empty()
                st.warning(
                    f"Not enough documents match the criteria. Found {len(filtered_df)}, need at least {min_cluster_size}."
                )
                return

            # Process content
            progress_bar.progress(0.4)
            status_text.text("Identifying themes...")

            processed_df = pd.DataFrame(
                {
                    "Content": filtered_df["Content"],
                    "Title": filtered_df["Title"],
                    "date_of_report": filtered_df["date_of_report"],
                    "URL": filtered_df["URL"],
                    "categories": filtered_df["categories"],
                }
            )

            progress_bar.progress(0.6)
            status_text.text("Analyzing patterns...")

            # Prepare vectorizer parameters
            vectorizer_params = {}
            if vectorizer_type == "weighted":
                vectorizer_params.update(
                    {"tf_scheme": tf_scheme, "idf_scheme": idf_scheme}
                )
            elif vectorizer_type == "bm25":
                vectorizer_params.update({"k1": k1, "b": b})

            # Store vectorization settings in session state
            st.session_state.vectorizer_type = vectorizer_type
            st.session_state.update(vectorizer_params)

            # Perform clustering
            cluster_results = perform_semantic_clustering(
                processed_df,
                min_cluster_size=min_cluster_size,
                max_features=max_features,
                min_df=2 / len(processed_df),
                max_df=0.95,
                similarity_threshold=0.3,
            )

            progress_bar.progress(0.8)
            status_text.text("Generating summaries...")

            # Store results
            st.session_state.topic_model = cluster_results

            progress_bar.progress(1.0)
            status_text.text("Analysis complete!")

            progress_bar.empty()
            status_text.empty()

            # Display results
            render_summary_tab(cluster_results, processed_df)

        except Exception as e:
            progress_bar.empty()
            status_text.empty()
            st.error(f"Analysis error: {str(e)}")
            logging.error(f"Analysis error: {e}", exc_info=True)


def render_topic_options():
    """Render enhanced topic analysis options in a clear layout"""

    st.subheader("Analysis Settings")

    # Create two columns for main settings
    col1, col2 = st.columns(2)

    with col1:
        st.markdown("##### Text Processing")
        vectorizer_type = st.selectbox(
            "Vectorization Method",
            options=["tfidf", "bm25", "weighted"],
            help="Choose how to convert text to numerical features:\n"
            + "- TF-IDF: Classic term frequency-inverse document frequency\n"
            + "- BM25: Enhanced version of TF-IDF used in search engines\n"
            + "- Weighted: Customizable term and document weighting",
        )

        # Show specific parameters based on vectorizer type
        if vectorizer_type == "bm25":
            st.markdown("##### BM25 Parameters")
            k1 = st.slider(
                "Term Saturation (k1)",
                min_value=0.5,
                max_value=3.0,
                value=1.5,
                step=0.1,
                help="Controls how quickly term frequency saturates (higher = slower)",
            )
            b = st.slider(
                "Length Normalization (b)",
                min_value=0.0,
                max_value=1.0,
                value=0.75,
                step=0.05,
                help="How much to penalize long documents",
            )

        elif vectorizer_type == "weighted":
            st.markdown("##### Weighting Schemes")
            tf_scheme = st.selectbox(
                "Term Frequency Scheme",
                options=["raw", "log", "binary", "augmented"],
                help="How to count term occurrences:\n"
                + "- Raw: Use actual counts\n"
                + "- Log: Logarithmic scaling\n"
                + "- Binary: Just presence/absence\n"
                + "- Augmented: Normalized frequency",
            )
            idf_scheme = st.selectbox(
                "Document Frequency Scheme",
                options=["smooth", "standard", "probabilistic"],
                help="How to weight document frequencies:\n"
                + "- Smooth: With smoothing factor\n"
                + "- Standard: Classic IDF\n"
                + "- Probabilistic: Based on probability",
            )

    with col2:
        st.markdown("##### Clustering Parameters")
        min_cluster_size = st.slider(
            "Minimum Cluster Size",
            min_value=2,
            max_value=10,
            value=3,
            help="Smallest allowed group of similar documents",
        )

        max_features = st.slider(
            "Maximum Features",
            min_value=1000,
            max_value=10000,
            value=5000,
            step=500,
            help="Maximum number of terms to consider",
        )

        min_similarity = st.slider(
            "Minimum Similarity",
            min_value=0.0,
            max_value=1.0,
            value=0.3,
            step=0.05,
            help="How similar documents must be to be grouped together",
        )

    # Advanced options in expander
    with st.expander("Advanced Settings"):
        st.markdown("##### Document Frequency Bounds")
        col3, col4 = st.columns(2)

        with col3:
            min_df = st.number_input(
                "Minimum Document Frequency",
                min_value=1,
                max_value=100,
                value=2,
                help="Minimum number of documents a term must appear in",
            )

        with col4:
            max_df = st.slider(
                "Maximum Document %",
                min_value=0.1,
                max_value=1.0,
                value=0.95,
                step=0.05,
                help="Maximum % of documents a term can appear in",
            )

        st.markdown("##### Visualization Settings")
        network_layout = st.selectbox(
            "Network Layout",
            options=["force", "circular", "random"],
            help="How to arrange nodes in topic networks",
        )

        show_weights = st.checkbox(
            "Show Edge Weights",
            value=True,
            help="Display connection strengths between terms",
        )

    return {
        "vectorizer_type": vectorizer_type,
        "vectorizer_params": {
            "k1": k1 if vectorizer_type == "bm25" else None,
            "b": b if vectorizer_type == "bm25" else None,
            "tf_scheme": tf_scheme if vectorizer_type == "weighted" else None,
            "idf_scheme": idf_scheme if vectorizer_type == "weighted" else None,
        },
        "cluster_params": {
            "min_cluster_size": min_cluster_size,
            "max_features": max_features,
            "min_similarity": min_similarity,
            "min_df": min_df,
            "max_df": max_df,
        },
        "viz_params": {"network_layout": network_layout, "show_weights": show_weights},
    }


def sort_reports(reports: List[Dict], order: str) -> List[Dict]:
    """Sort reports based on specified order"""
    if order == "date_desc":
        return sorted(
            reports,
            key=lambda x: datetime.strptime(x["date"], "%Y-%m-%d"),
            reverse=True,
        )
    elif order == "date_asc":
        return sorted(reports, key=lambda x: datetime.strptime(x["date"], "%Y-%m-%d"))
    return reports


def plot_category_distribution(df: pd.DataFrame) -> None:
    """Plot category distribution"""
    all_cats = []
    for cats in df["categories"].dropna():
        if isinstance(cats, list):
            all_cats.extend(cats)

    cat_counts = pd.Series(all_cats).value_counts()

    fig = px.bar(
        x=cat_counts.index,
        y=cat_counts.values,
        title="Category Distribution",
        labels={"x": "Category", "y": "Count"},
    )

    fig.update_layout(
        xaxis_title="Category", yaxis_title="Number of Reports", xaxis={"tickangle": 45}
    )

    st.plotly_chart(fig, use_container_width=True)


def plot_coroner_areas(df: pd.DataFrame) -> None:
    """Plot coroner areas distribution"""
    area_counts = df["coroner_area"].value_counts().head(20)

    fig = px.bar(
        x=area_counts.index,
        y=area_counts.values,
        title="Top 20 Coroner Areas",
        labels={"x": "Area", "y": "Count"},
    )

    fig.update_layout(
        xaxis_title="Coroner Area",
        yaxis_title="Number of Reports",
        xaxis={"tickangle": 45},
    )

    st.plotly_chart(fig, use_container_width=True)


def analyze_data_quality(df: pd.DataFrame) -> None:
    """Analyze and display data quality metrics for PFD reports"""

    # Calculate completeness metrics
    total_records = len(df)

    def calculate_completeness(field):
        if field not in df.columns:
            return 0
        non_empty = df[field].notna()
        if field == "categories":
            non_empty = df[field].apply(lambda x: isinstance(x, list) and len(x) > 0)
        return (non_empty.sum() / total_records) * 100

    completeness_metrics = {
        "Title": calculate_completeness("Title"),
        "Content": calculate_completeness("Content"),
        "Date of Report": calculate_completeness("date_of_report"),
        "Deceased Name": calculate_completeness("deceased_name"),
        "Coroner Name": calculate_completeness("coroner_name"),
        "Coroner Area": calculate_completeness("coroner_area"),
        "Categories": calculate_completeness("categories"),
    }

    # Calculate consistency metrics
    consistency_metrics = {
        "Title Format": (df["Title"].str.len() >= 10).mean() * 100,
        "Content Length": (df["Content"].str.len() >= 100).mean() * 100,
        "Date Format": df["date_of_report"].notna().mean() * 100,
        "Categories Format": df["categories"]
        .apply(lambda x: isinstance(x, list))
        .mean()
        * 100,
    }

    # Calculate PDF metrics
    pdf_columns = [
        col for col in df.columns if col.startswith("PDF_") and col.endswith("_Path")
    ]
    reports_with_pdf = df[pdf_columns].notna().any(axis=1).sum()
    reports_with_multiple_pdfs = (df[pdf_columns].notna().sum(axis=1) > 1).sum()

    pdf_metrics = {
        "Reports with PDFs": (reports_with_pdf / total_records) * 100,
        "Reports with Multiple PDFs": (reports_with_multiple_pdfs / total_records)
        * 100,
    }

    # Display metrics using Streamlit
    st.subheader("Data Quality Analysis")

    # Completeness
    st.markdown("### Field Completeness")
    completeness_df = pd.DataFrame(
        list(completeness_metrics.items()), columns=["Field", "Completeness %"]
    )
    fig_completeness = px.bar(
        completeness_df,
        x="Field",
        y="Completeness %",
        title="Field Completeness Analysis",
    )
    fig_completeness.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_completeness, use_container_width=True)

    # Consistency
    st.markdown("### Data Consistency")
    consistency_df = pd.DataFrame(
        list(consistency_metrics.items()), columns=["Metric", "Consistency %"]
    )
    fig_consistency = px.bar(
        consistency_df, x="Metric", y="Consistency %", title="Data Consistency Analysis"
    )
    fig_consistency.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_consistency, use_container_width=True)

    # PDF Analysis
    st.markdown("### PDF Attachment Analysis")
    pdf_df = pd.DataFrame(list(pdf_metrics.items()), columns=["Metric", "Percentage"])
    fig_pdf = px.bar(pdf_df, x="Metric", y="Percentage", title="PDF Coverage Analysis")
    fig_pdf.update_layout(xaxis_tickangle=-45)
    st.plotly_chart(fig_pdf, use_container_width=True)

    # Summary metrics
    col1, col2, col3 = st.columns(3)

    with col1:
        st.metric(
            "Average Completeness",
            f"{np.mean(list(completeness_metrics.values())):.1f}%",
        )

    with col2:
        st.metric(
            "Average Consistency", f"{np.mean(list(consistency_metrics.values())):.1f}%"
        )

    with col3:
        st.metric("PDF Coverage", f"{pdf_metrics['Reports with PDFs']:.1f}%")

    # Detailed quality issues
    st.markdown("### Detailed Quality Issues")

    issues = []

    # Check for missing crucial fields
    for field, completeness in completeness_metrics.items():
        if completeness < 95:  # Less than 95% complete
            issues.append(
                f"- {field} is {completeness:.1f}% complete ({total_records - int(completeness * total_records / 100)} records missing)"
            )

    # Check for consistency issues
    for metric, consistency in consistency_metrics.items():
        if consistency < 90:  # Less than 90% consistent
            issues.append(f"- {metric} shows {consistency:.1f}% consistency")

    # Check PDF coverage
    if pdf_metrics["Reports with PDFs"] < 90:
        issues.append(
            f"- {100 - pdf_metrics['Reports with PDFs']:.1f}% of reports lack PDF attachments"
        )

    if issues:
        st.markdown("The following quality issues were identified:")
        for issue in issues:
            st.markdown(issue)
    else:
        st.success("No significant quality issues found in the dataset.")


def display_topic_network(lda, feature_names):
    """Display word similarity network with interactive filters"""
    # st.markdown("### Word Similarity Network")
    st.markdown(
        "This network shows relationships between words based on their co-occurrence in documents."
    )

    # Store base network data in session state if not already present
    if "network_data" not in st.session_state:
        # Get word counts across all documents
        word_counts = lda.components_.sum(axis=0)
        top_word_indices = word_counts.argsort()[
            : -100 - 1 : -1
        ]  # Store more words initially

        # Create word co-occurrence matrix
        word_vectors = normalize(lda.components_.T[top_word_indices])
        word_similarities = cosine_similarity(word_vectors)

        st.session_state.network_data = {
            "word_counts": word_counts,
            "top_word_indices": top_word_indices,
            "word_similarities": word_similarities,
            "feature_names": feature_names,
        }

    # Network filters with keys to prevent rerun
    col1, col2, col3 = st.columns(3)
    with col1:
        min_similarity = st.slider(
            "Minimum Similarity",
            min_value=0.0,
            max_value=1.0,
            value=0.9,
            step=0.05,
            help="Higher values show stronger connections only",
            key="network_min_similarity",
        )
    with col2:
        max_words = st.slider(
            "Number of Words",
            min_value=10,
            max_value=100,
            value=30,
            step=5,
            help="Number of most frequent words to show",
            key="network_max_words",
        )
    with col3:
        min_connections = st.slider(
            "Minimum Connections",
            min_value=1,
            max_value=10,
            value=5,
            help="Minimum number of connections per word",
            key="network_min_connections",
        )

    # Create network graph based on current filters
    G = nx.Graph()

    # Get stored data
    word_counts = st.session_state.network_data["word_counts"]
    word_similarities = st.session_state.network_data["word_similarities"]
    top_word_indices = st.session_state.network_data["top_word_indices"][:max_words]
    feature_names = st.session_state.network_data["feature_names"]

    # Add nodes
    for idx, word_idx in enumerate(top_word_indices):
        G.add_node(idx, name=feature_names[word_idx], freq=float(word_counts[word_idx]))

    # Add edges based on current similarity threshold
    for i in range(len(top_word_indices)):
        for j in range(i + 1, len(top_word_indices)):
            similarity = word_similarities[i, j]
            if similarity > min_similarity:
                G.add_edge(i, j, weight=float(similarity))

    # Filter nodes by minimum connections
    nodes_to_remove = []
    for node in G.nodes():
        if G.degree(node) < min_connections:
            nodes_to_remove.append(node)
    G.remove_nodes_from(nodes_to_remove)

    if len(G.nodes()) == 0:
        st.warning(
            "No nodes match the current filter criteria. Try adjusting the filters."
        )
        return

    # Create visualization
    pos = nx.spring_layout(G, k=1 / np.sqrt(len(G.nodes())), iterations=50)

    # Create edge traces with varying thickness and color based on weight
    edge_traces = []
    for edge in G.edges(data=True):
        x0, y0 = pos[edge[0]]
        x1, y1 = pos[edge[1]]
        weight = edge[2]["weight"]

        edge_trace = go.Scatter(
            x=[x0, x1, None],
            y=[y0, y1, None],
            line=dict(width=weight * 3, color=f"rgba(100,100,100,{weight})"),
            hoverinfo="none",
            mode="lines",
        )
        edge_traces.append(edge_trace)

    # Create node trace with size based on frequency
    node_x = []
    node_y = []
    node_text = []
    node_size = []

    for node in G.nodes():
        x, y = pos[node]
        node_x.append(x)
        node_y.append(y)
        freq = G.nodes[node]["freq"]
        name = G.nodes[node]["name"]
        connections = G.degree(node)
        node_text.append(
            f"{name}<br>Frequency: {freq:.0f}<br>Connections: {connections}"
        )
        node_size.append(np.sqrt(freq) * 10)

    node_trace = go.Scatter(
        x=node_x,
        y=node_y,
        mode="markers+text",
        hoverinfo="text",
        text=node_text,
        textposition="top center",
        marker=dict(
            size=node_size, line=dict(width=1), color="lightblue", sizemode="area"
        ),
    )

    # Create figure
    fig = go.Figure(
        data=edge_traces + [node_trace],
        layout=go.Layout(
            title=f"Word Network ({len(G.nodes())} words, {len(G.edges())} connections)",
            titlefont_size=16,
            showlegend=False,
            hovermode="closest",
            margin=dict(b=20, l=5, r=5, t=40),
            xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
            yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        ),
    )

    st.plotly_chart(fig, use_container_width=True)

    # Add network statistics
    st.markdown("### Network Statistics")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("Number of Words", len(G.nodes()))
    with col2:
        st.metric("Number of Connections", len(G.edges()))
    with col3:
        if len(G.nodes()) > 0:
            density = 2 * len(G.edges()) / (len(G.nodes()) * (len(G.nodes()) - 1))
            st.metric("Network Density", f"{density:.2%}")


def get_top_words(model, feature_names, topic_idx, n_words=10):
    """Get top words for a given topic"""
    return [
        feature_names[i]
        for i in model.components_[topic_idx].argsort()[: -n_words - 1 : -1]
    ]


def render_file_upload():
    """Render file upload section"""
    st.header("Upload Existing Data")

    # Generate unique key for the file uploader
    upload_key = f"file_uploader_{int(time.time() * 1000)}"

    uploaded_file = st.file_uploader(
        "Upload CSV or Excel file", type=["csv", "xlsx"], key=upload_key
    )

    if uploaded_file is not None:
        try:
            if uploaded_file.name.endswith(".csv"):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)

            # Process uploaded data
            df = process_scraped_data(df)

            # Clear any existing data first
            st.session_state.current_data = None
            st.session_state.scraped_data = None
            st.session_state.uploaded_data = None
            st.session_state.data_source = None

            # Then set new data
            st.session_state.uploaded_data = df.copy()
            st.session_state.data_source = "uploaded"
            st.session_state.current_data = df.copy()

            st.success("File uploaded and processed successfully!")

            # Show the uploaded data
            st.subheader("Uploaded Data Preview")
            st.dataframe(
                df,
                column_config={
                    "URL": st.column_config.LinkColumn("Report Link"),
                    "date_of_report": st.column_config.DateColumn("Date of Report"),
                    "categories": st.column_config.ListColumn("Categories"),
                },
                hide_index=True,
            )

            return True

        except Exception as e:
            st.error(f"Error uploading file: {str(e)}")
            logging.error(f"File upload error: {e}", exc_info=True)
            return False

    return False


def initialize_session_state():
    """Initialize all required session state variables"""
    # Initialize basic state variables if they don't exist
    if not hasattr(st.session_state, "initialized"):
        # Clear all existing session state
        for key in list(st.session_state.keys()):
            del st.session_state[key]

        # Set new session state variables
        st.session_state.data_source = None
        st.session_state.current_data = None
        st.session_state.scraped_data = None
        st.session_state.uploaded_data = None
        st.session_state.topic_model = None
        st.session_state.cleanup_done = False
        st.session_state.last_scrape_time = None
        st.session_state.last_upload_time = None
        st.session_state.analysis_filters = {
            "date_range": None,
            "selected_categories": None,
            "selected_areas": None,
        }
        st.session_state.topic_model_settings = {
            "num_topics": 5,
            "max_features": 1000,
            "similarity_threshold": 0.3,
        }
        st.session_state.initialized = True

    # Perform PDF cleanup if not done
    if not st.session_state.cleanup_done:
        try:
            pdf_dir = "pdfs"
            os.makedirs(pdf_dir, exist_ok=True)

            current_time = time.time()
            cleanup_count = 0

            for file in os.listdir(pdf_dir):
                file_path = os.path.join(pdf_dir, file)
                try:
                    if os.path.isfile(file_path):
                        if os.stat(file_path).st_mtime < current_time - 86400:
                            os.remove(file_path)
                            cleanup_count += 1
                except Exception as e:
                    logging.warning(f"Error cleaning up file {file_path}: {e}")
                    continue

            if cleanup_count > 0:
                logging.info(f"Cleaned up {cleanup_count} old PDF files")
        except Exception as e:
            logging.error(f"Error during PDF cleanup: {e}")
        finally:
            st.session_state.cleanup_done = True


def validate_data(data: pd.DataFrame, purpose: str = "analysis") -> Tuple[bool, str]:
    """
    Validate data for different purposes

    Args:
        data: DataFrame to validate
        purpose: Purpose of validation ('analysis' or 'topic_modeling')

    Returns:
        tuple: (is_valid, message)
    """
    if data is None:
        return False, "No data available. Please scrape or upload data first."

    if not isinstance(data, pd.DataFrame):
        return False, "Invalid data format. Expected pandas DataFrame."

    if len(data) == 0:
        return False, "Dataset is empty."

    if purpose == "analysis":
        required_columns = ["date_of_report", "categories", "coroner_area"]
        missing_columns = [col for col in required_columns if col not in data.columns]
        if missing_columns:
            return False, f"Missing required columns: {', '.join(missing_columns)}"

    elif purpose == "topic_modeling":
        if "Content" not in data.columns:
            return False, "Missing required column: Content"

        valid_docs = data["Content"].dropna().str.strip().str.len() > 0
        if valid_docs.sum() < 2:
            return (
                False,
                "Not enough valid documents found. Please ensure you have documents with text content.",
            )

    # Add type checking for critical columns
    if "date_of_report" in data.columns and not pd.api.types.is_datetime64_any_dtype(
        data["date_of_report"]
    ):
        try:
            pd.to_datetime(data["date_of_report"])
        except Exception:
            return False, "Invalid date format in date_of_report column."

    if "categories" in data.columns:
        if (
            not data["categories"]
            .apply(lambda x: isinstance(x, (list, type(None))))
            .all()
        ):
            return False, "Categories must be stored as lists or None values."

    return True, "Data is valid"


def is_response(row: pd.Series) -> bool:
    """
    Check if a report is a response document based on its metadata and content

    Args:
        row: DataFrame row containing report data

    Returns:
        bool: True if document is a response, False otherwise
    """
    try:
        # Check PDF names for response indicators
        pdf_response = False
        for i in range(1, 10):  # Check PDF_1 to PDF_9
            pdf_name = str(row.get(f"PDF_{i}_Name", "")).lower()
            if "response" in pdf_name or "reply" in pdf_name:
                pdf_response = True
                break

        # Check title for response indicators
        title = str(row.get("Title", "")).lower()
        title_response = any(
            word in title for word in ["response", "reply", "answered"]
        )

        # Check content for response indicators
        content = str(row.get("Content", "")).lower()
        content_response = any(
            phrase in content
            for phrase in [
                "in response to",
                "responding to",
                "reply to",
                "response to",
                "following the regulation 28",
            ]
        )

        return pdf_response or title_response or content_response

    except Exception as e:
        logging.error(f"Error checking response type: {e}")
        return False


def plot_timeline(df: pd.DataFrame) -> None:
    """Plot timeline of reports with improved formatting"""
    timeline_data = (
        df.groupby(pd.Grouper(key="date_of_report", freq="M")).size().reset_index()
    )
    timeline_data.columns = ["Date", "Count"]

    fig = px.line(
        timeline_data,
        x="Date",
        y="Count",
        title="Reports Timeline",
        labels={"Count": "Number of Reports"},
    )

    fig.update_layout(
        xaxis_title="Date",
        yaxis_title="Number of Reports",
        hovermode="x unified",
        yaxis=dict(
            tickmode="linear",
            tick0=0,
            dtick=1,  # Integer steps
            rangemode="nonnegative",  # Ensure y-axis starts at 0 or above
        ),
        xaxis=dict(tickformat="%B %Y", tickangle=45),  # Month Year format
    )

    st.plotly_chart(fig, use_container_width=True)


def plot_monthly_distribution(df: pd.DataFrame) -> None:
    """Plot monthly distribution with improved formatting"""
    # Format dates as Month Year
    df["month_year"] = df["date_of_report"].dt.strftime("%B %Y")
    monthly_counts = df["month_year"].value_counts().sort_index()

    fig = px.bar(
        x=monthly_counts.index,
        y=monthly_counts.values,
        labels={"x": "Month", "y": "Number of Reports"},
        title="Monthly Distribution of Reports",
    )

    fig.update_layout(
        xaxis_title="Month",
        yaxis_title="Number of Reports",
        xaxis_tickangle=45,
        yaxis=dict(
            tickmode="linear",
            tick0=0,
            dtick=1,  # Integer steps
            rangemode="nonnegative",
        ),
        bargap=0.2,
    )

    st.plotly_chart(fig, use_container_width=True)


def plot_yearly_comparison(df: pd.DataFrame) -> None:
    """Plot year-over-year comparison with improved formatting"""
    yearly_counts = df["date_of_report"].dt.year.value_counts().sort_index()

    fig = px.line(
        x=yearly_counts.index.astype(int),  # Convert to integer years
        y=yearly_counts.values,
        markers=True,
        labels={"x": "Year", "y": "Number of Reports"},
        title="Year-over-Year Report Volumes",
    )

    # Calculate appropriate y-axis maximum
    max_count = yearly_counts.max()
    y_max = max_count + (1 if max_count < 10 else 2)  # Add some padding

    fig.update_layout(
        xaxis_title="Year",
        yaxis_title="Number of Reports",
        xaxis=dict(
            tickmode="linear",
            tick0=yearly_counts.index.min(),
            dtick=1,  # Show every year
            tickformat="d",  # Format as integer
        ),
        yaxis=dict(
            tickmode="linear", tick0=0, dtick=1, range=[0, y_max]  # Integer steps
        ),
    )

    st.plotly_chart(fig, use_container_width=True)


def export_to_excel(df: pd.DataFrame) -> bytes:
    """
    Export DataFrame to Excel bytes with proper formatting
    """
    try:
        if df is None or len(df) == 0:
            raise ValueError("No data available to export")

        # Create clean copy for export
        df_export = df.copy()

        # Format dates to UK format
        if "date_of_report" in df_export.columns:
            df_export["date_of_report"] = df_export["date_of_report"].dt.strftime(
                "%d/%m/%Y"
            )

        # Handle list columns (like categories)
        for col in df_export.columns:
            if df_export[col].dtype == "object":
                df_export[col] = df_export[col].apply(
                    lambda x: ", ".join(x)
                    if isinstance(x, list)
                    else str(x)
                    if pd.notna(x)
                    else ""
                )

        # Create output buffer
        output = io.BytesIO()

        # Write to Excel
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_export.to_excel(writer, sheet_name="Reports", index=False)

            # Get the worksheet
            worksheet = writer.sheets["Reports"]

            # Auto-adjust column widths
            for idx, col in enumerate(df_export.columns, 1):
                max_length = max(
                    df_export[col].astype(str).apply(len).max(), len(str(col))
                )
                adjusted_width = min(max_length + 2, 50)
                column_letter = get_column_letter(
                    idx
                )  # Use openpyxl's get_column_letter
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Add filters to header row
            worksheet.auto_filter.ref = worksheet.dimensions

            # Freeze the header row
            worksheet.freeze_panes = "A2"

        # Get the bytes value
        output.seek(0)
        return output.getvalue()

    except Exception as e:
        logging.error(f"Error exporting to Excel: {e}", exc_info=True)
        raise Exception(f"Failed to export data to Excel: {str(e)}")


def show_export_options(df: pd.DataFrame, prefix: str):
    """Show export options for the data with descriptive filename and unique keys"""
    try:
        st.subheader("Export Options")

        # Generate timestamp and random suffix for unique keys
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        random_suffix = "".join(
            random.choices(string.ascii_lowercase + string.digits, k=6)
        )
        unique_id = f"{timestamp}_{random_suffix}"
        filename = f"pfd_reports_{prefix}_{timestamp}"

        col1, col2 = st.columns(2)

        # CSV Export
        with col1:
            try:
                # Create export copy with formatted dates
                df_csv = df.copy()
                if "date_of_report" in df_csv.columns:
                    df_csv["date_of_report"] = df_csv["date_of_report"].dt.strftime(
                        "%d/%m/%Y"
                    )

                csv = df_csv.to_csv(index=False).encode("utf-8")
                st.download_button(
                    "📥 Download Reports (CSV)",
                    csv,
                    f"{filename}.csv",
                    "text/csv",
                    key=f"download_csv_{prefix}_{unique_id}",
                )
            except Exception as e:
                st.error(f"Error preparing CSV export: {str(e)}")

        # Excel Export
        with col2:
            try:
                excel_data = export_to_excel(df)
                st.download_button(
                    "📥 Download Reports (Excel)",
                    excel_data,
                    f"{filename}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_excel_{prefix}_{unique_id}",
                )
            except Exception as e:
                st.error(f"Error preparing Excel export: {str(e)}")

        # PDF Download
        if any(col.startswith("PDF_") and col.endswith("_Path") for col in df.columns):
            st.subheader("Download PDFs")
            if st.button(f"Download all PDFs", key=f"pdf_button_{prefix}_{unique_id}"):
                with st.spinner("Preparing PDF download..."):
                    try:
                        pdf_zip_path = f"{filename}_pdfs.zip"

                        with zipfile.ZipFile(pdf_zip_path, "w") as zipf:
                            pdf_columns = [
                                col
                                for col in df.columns
                                if col.startswith("PDF_") and col.endswith("_Path")
                            ]
                            added_files = set()

                            for col in pdf_columns:
                                paths = df[col].dropna()
                                for pdf_path in paths:
                                    if (
                                        pdf_path
                                        and os.path.exists(pdf_path)
                                        and pdf_path not in added_files
                                    ):
                                        zipf.write(pdf_path, os.path.basename(pdf_path))
                                        added_files.add(pdf_path)

                        with open(pdf_zip_path, "rb") as f:
                            st.download_button(
                                "📦 Download All PDFs (ZIP)",
                                f.read(),
                                pdf_zip_path,
                                "application/zip",
                                key=f"download_pdfs_zip_{prefix}_{unique_id}",
                            )

                        # Cleanup zip file
                        os.remove(pdf_zip_path)
                    except Exception as e:
                        st.error(f"Error preparing PDF download: {str(e)}")

    except Exception as e:
        st.error(f"Error setting up export options: {str(e)}")
        logging.error(f"Export options error: {e}", exc_info=True)


def extract_advanced_topics(
    data: pd.DataFrame,
    num_topics: int = 5,
    max_features: int = 1000,
    min_df: int = 2,
    n_iterations: int = 20,
    min_similarity: float = 0.9,
) -> Tuple[LatentDirichletAllocation, CountVectorizer, np.ndarray]:
    """
    Advanced topic modeling with comprehensive preprocessing and error handling

    Args:
        data (pd.DataFrame): Input DataFrame containing documents
        num_topics (int): Number of topics to extract
        max_features (int): Maximum number of features to use
        min_df (int): Minimum document frequency for terms
        n_iterations (int): Maximum number of iterations for LDA
        min_similarity (float): Minimum similarity threshold for the word similarity network

    Returns:
        Tuple containing LDA model, vectorizer, and document-topic distribution
    """
    try:
        # Extensive logging
        logging.info(f"Starting topic modeling with {len(data)} documents")
        logging.info(
            f"Parameters: topics={num_topics}, max_features={max_features}, min_df={min_df}, min_similarity={min_similarity}"
        )

        # Validate input data
        if data is None or len(data) == 0:
            raise ValueError("No data provided for topic modeling")

        # Remove duplicate documents based on content
        def prepare_document(doc: str) -> str:
            """Clean and prepare individual documents"""
            if pd.isna(doc):
                return None

            # Aggressive text cleaning
            cleaned_doc = clean_text_for_modeling(str(doc))

            # Minimum length check
            return cleaned_doc if len(cleaned_doc.split()) > 3 else None

        # Process documents
        documents = data["Content"].apply(prepare_document).dropna().unique().tolist()

        logging.info(f"Processed {len(documents)} unique valid documents")

        # Validate document count
        if len(documents) < num_topics:
            adjusted_topics = max(2, len(documents) // 2)
            logging.warning(
                f"Not enough documents for {num_topics} topics. Adjusting to {adjusted_topics}"
            )
            num_topics = adjusted_topics

        # Vectorization with robust settings
        vectorizer = CountVectorizer(
            max_features=max_features,
            min_df=min(min_df, max(2, len(documents) // 10)),  # Adaptive min_df
            max_df=0.95,
            stop_words="english",
        )

        # Create document-term matrix
        dtm = vectorizer.fit_transform(documents)
        feature_names = vectorizer.get_feature_names_out()

        logging.info(f"Document-term matrix shape: {dtm.shape}")
        logging.info(f"Number of features: {len(feature_names)}")

        # LDA with robust parameters
        lda_model = LatentDirichletAllocation(
            n_components=num_topics,
            random_state=42,
            learning_method="online",
            learning_offset=50.0,
            max_iter=n_iterations,
            doc_topic_prior=None,  # Let scikit-learn auto-estimate
            topic_word_prior=None,  # Let scikit-learn auto-estimate
        )

        # Fit LDA model
        doc_topics = lda_model.fit_transform(dtm)

        # Add logging of results
        logging.info("Topic modeling completed successfully")
        logging.info(f"Document-topic matrix shape: {doc_topics.shape}")

        return lda_model, vectorizer, doc_topics

    except Exception as e:
        logging.error(f"Topic modeling failed: {e}", exc_info=True)
        raise


def is_response(row: pd.Series) -> bool:
    """
    Check if a document is a response based on its metadata and content
    """
    try:
        # Check PDF types first (most reliable)
        for i in range(1, 5):  # Check PDF_1 to PDF_4
            pdf_type = str(row.get(f"PDF_{i}_Type", "")).lower()
            if pdf_type == "response":
                return True

        # Check PDF names as backup
        for i in range(1, 5):
            pdf_name = str(row.get(f"PDF_{i}_Name", "")).lower()
            if "response" in pdf_name or "reply" in pdf_name:
                return True

        # Check title and content as final fallback
        title = str(row.get("Title", "")).lower()
        if any(word in title for word in ["response", "reply", "answered"]):
            return True

        content = str(row.get("Content", "")).lower()
        return any(
            phrase in content
            for phrase in [
                "in response to",
                "responding to",
                "reply to",
                "response to",
                "following the regulation 28",
                "following receipt of the regulation 28",
            ]
        )

    except Exception as e:
        logging.error(f"Error checking response type: {e}")
        return False


def normalize_category(category: str) -> str:
    """Normalize category string for consistent matching"""
    if not category:
        return ""
    # Convert to lowercase and remove extra whitespace
    normalized = " ".join(str(category).lower().split())
    # Remove common separators and special characters
    normalized = re.sub(r"[,;|•·⋅‣⁃▪▫–—-]", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def match_category(category: str, standard_categories: List[str]) -> Optional[str]:
    """Match a category against standard categories with fuzzy matching"""
    if not category:
        return None

    normalized_category = normalize_category(category)
    normalized_standards = {normalize_category(cat): cat for cat in standard_categories}

    # Try exact match first
    if normalized_category in normalized_standards:
        return normalized_standards[normalized_category]

    # Try partial matching
    for norm_std, original_std in normalized_standards.items():
        if normalized_category in norm_std or norm_std in normalized_category:
            return original_std

    return category  # Return original if no match found


def extract_categories(category_text: str, standard_categories: List[str]) -> List[str]:
    """Extract and normalize categories from raw text"""
    if not category_text:
        return []

    # Replace common separators with a standard one
    category_text = re.sub(r"\s*[|,;]\s*", "|", category_text)
    category_text = re.sub(r"[•·⋅‣⁃▪▫–—-]\s*", "|", category_text)
    category_text = re.sub(r"\s{2,}", "|", category_text)
    category_text = re.sub(r"\n+", "|", category_text)

    # Split and clean categories
    raw_categories = category_text.split("|")
    cleaned_categories = []

    for cat in raw_categories:
        cleaned_cat = clean_text(cat).strip()
        if cleaned_cat and not re.match(r"^[\s|,;]+$", cleaned_cat):
            matched_cat = match_category(cleaned_cat, standard_categories)
            if matched_cat:
                cleaned_categories.append(matched_cat)

    # Remove duplicates while preserving order
    seen = set()
    return [
        x
        for x in cleaned_categories
        if not (normalize_category(x) in seen or seen.add(normalize_category(x)))
    ]


def filter_by_categories(
    df: pd.DataFrame, selected_categories: List[str]
) -> pd.DataFrame:
    """
    Filter DataFrame by categories with fuzzy matching

    Args:
        df: DataFrame containing 'categories' column
        selected_categories: List of categories to filter by

    Returns:
        Filtered DataFrame
    """
    if not selected_categories:
        return df

    def has_matching_category(row_categories):
        if not isinstance(row_categories, list):
            return False

        # Normalize categories for comparison
        row_cats_norm = [cat.lower().strip() for cat in row_categories if cat]
        selected_cats_norm = [cat.lower().strip() for cat in selected_categories if cat]

        for row_cat in row_cats_norm:
            for selected_cat in selected_cats_norm:
                # Check for partial matches in either direction
                if row_cat in selected_cat or selected_cat in row_cat:
                    return True
        return False

    return df[df["categories"].apply(has_matching_category)]


def filter_by_document_type(df: pd.DataFrame, doc_types: List[str]) -> pd.DataFrame:
    """
    Filter DataFrame based on document types
    """
    if not doc_types:
        return df

    filtered_df = df.copy()
    is_response_mask = filtered_df.apply(is_response, axis=1)

    if len(doc_types) == 1:
        if "Response" in doc_types:
            return filtered_df[is_response_mask]
        elif "Report" in doc_types:
            return filtered_df[~is_response_mask]

    return filtered_df


def extract_topic_insights(lda_model, vectorizer, doc_topics, data: pd.DataFrame):
    """Extract insights from topic modeling results with improved error handling"""
    try:
        # Get feature names and initialize results
        feature_names = vectorizer.get_feature_names_out()
        topics_data = []

        # Ensure we have valid data
        valid_data = data[data["Content"].notna()].copy()
        if len(valid_data) == 0:
            raise ValueError("No valid documents found in dataset")

        # Calculate document frequencies with error handling
        doc_freq = {}
        for doc in valid_data["Content"]:
            try:
                words = set(clean_text_for_modeling(str(doc)).split())
                for word in words:
                    doc_freq[word] = doc_freq.get(word, 0) + 1
            except Exception as e:
                logging.warning(f"Error processing document: {str(e)}")
                continue

        # Process each topic
        for idx, topic in enumerate(lda_model.components_):
            try:
                # Get top words
                top_word_indices = topic.argsort()[: -50 - 1 : -1]
                topic_words = []

                for i in top_word_indices:
                    word = feature_names[i]
                    if len(word) > 1:
                        weight = float(topic[i])
                        topic_words.append(
                            {
                                "word": word,
                                "weight": weight,
                                "count": doc_freq.get(word, 0),
                                "documents": doc_freq.get(word, 0),
                            }
                        )

                # Get representative documents
                doc_scores = doc_topics[:, idx]
                top_doc_indices = doc_scores.argsort()[:-11:-1]

                related_docs = []
                for doc_idx in top_doc_indices:
                    if doc_scores[doc_idx] > 0.01:  # At least 1% relevance
                        if doc_idx < len(valid_data):
                            doc_row = valid_data.iloc[doc_idx]
                            doc_content = str(doc_row.get("Content", ""))

                            related_docs.append(
                                {
                                    "title": doc_row.get("Title", ""),
                                    "date": doc_row.get("date_of_report", ""),
                                    "relevance": float(doc_scores[doc_idx]),
                                    "summary": doc_content[:300] + "..."
                                    if len(doc_content) > 300
                                    else doc_content,
                                }
                            )

                # Generate topic description
                meaningful_words = [word["word"] for word in topic_words[:5]]
                label = " & ".join(meaningful_words[:3]).title()

                topic_data = {
                    "id": idx,
                    "label": label,
                    "description": f"Topic frequently mentions: {', '.join(meaningful_words)}",
                    "words": topic_words,
                    "representativeDocs": related_docs,
                    "prevalence": round((doc_scores > 0.05).mean() * 100, 1),
                }

                topics_data.append(topic_data)

            except Exception as e:
                logging.error(f"Error processing topic {idx}: {str(e)}")
                continue

        if not topics_data:
            raise ValueError("No valid topics could be extracted")

        return topics_data

    except Exception as e:
        logging.error(f"Error extracting topic insights: {str(e)}", exc_info=True)
        raise Exception(f"Failed to extract topic insights: {str(e)}")


def display_topic_analysis(topics_data):
    """Display topic analysis results"""
    for topic in topics_data:
        st.markdown(f"## Topic {topic['id'] + 1}: {topic['label']}")
        st.markdown(f"**Prevalence:** {topic['prevalence']}% of documents")
        st.markdown(f"**Description:** {topic['description']}")

        # Display key terms
        st.markdown("### Key Terms")
        terms_data = pd.DataFrame(topic["words"])
        if not terms_data.empty:
            st.dataframe(
                terms_data,
                column_config={
                    "word": st.column_config.TextColumn("Term"),
                    "weight": st.column_config.NumberColumn("Weight", format="%.4f"),
                    "count": st.column_config.NumberColumn("Document Count"),
                },
                hide_index=True,
            )

        # Display representative documents
        st.markdown("### Representative Documents")
        for doc in topic["representativeDocs"]:
            with st.expander(f"{doc['title']} (Relevance: {doc['relevance']:.2%})"):
                st.markdown(f"**Date:** {doc['date']}")
                st.markdown(doc["summary"])

        st.markdown("---")


# Initialize NLTK resources
def initialize_nltk():
    """Initialize required NLTK resources with error handling"""
    try:
        resources = ["punkt", "stopwords", "averaged_perceptron_tagger"]
        for resource in resources:
            try:
                if resource == "punkt":
                    nltk.data.find("tokenizers/punkt")
                elif resource == "stopwords":
                    nltk.data.find("corpora/stopwords")
                elif resource == "averaged_perceptron_tagger":
                    nltk.data.find("taggers/averaged_perceptron_tagger")
            except LookupError:
                nltk.download(resource)
    except Exception as e:
        logging.error(f"Error initializing NLTK resources: {e}")
        raise


def perform_semantic_clustering(
    data: pd.DataFrame,
    min_cluster_size: int = 3,
    max_features: int = 5000,
    min_df: float = 0.01,
    max_df: float = 0.95,
    similarity_threshold: float = 0.3,
) -> Dict:
    """
    Perform semantic clustering with improved cluster selection
    """
    try:
        # Initialize NLTK resources
        initialize_nltk()

        # Validate input data
        if "Content" not in data.columns:
            raise ValueError("Input data must contain 'Content' column")

        processed_texts = data["Content"].apply(clean_text_for_modeling)
        valid_mask = processed_texts.notna() & (processed_texts != "")
        processed_texts = processed_texts[valid_mask]

        if len(processed_texts) == 0:
            raise ValueError("No valid text content found after preprocessing")

        # Keep the original data for display
        display_data = data[valid_mask].copy()

        # Calculate optimal parameters based on dataset size
        n_docs = len(processed_texts)
        min_clusters = max(2, min(3, n_docs // 20))  # More conservative minimum
        max_clusters = max(3, min(8, n_docs // 10))  # More conservative maximum

        # Get vectorization parameters from session state
        vectorizer_type = st.session_state.get("vectorizer_type", "tfidf")
        vectorizer_params = {}

        if vectorizer_type == "bm25":
            vectorizer_params.update(
                {
                    "k1": st.session_state.get("bm25_k1", 1.5),
                    "b": st.session_state.get("bm25_b", 0.75),
                }
            )
        elif vectorizer_type == "weighted":
            vectorizer_params.update(
                {
                    "tf_scheme": st.session_state.get("tf_scheme", "raw"),
                    "idf_scheme": st.session_state.get("idf_scheme", "smooth"),
                }
            )

        # Create the vectorizer
        vectorizer = get_vectorizer(
            vectorizer_type=vectorizer_type,
            max_features=max_features,
            min_df=max(min_df, 3 / len(processed_texts)),
            max_df=min(max_df, 0.7),
            **vectorizer_params,
        )

        # Create document vectors
        tfidf_matrix = vectorizer.fit_transform(processed_texts)
        feature_names = vectorizer.get_feature_names_out()

        # Find optimal number of clusters
        best_n_clusters, best_labels = find_optimal_clusters(
            tfidf_matrix,
            min_clusters=min_clusters,
            max_clusters=max_clusters,
            min_cluster_size=min_cluster_size,
        )

        # Calculate final clustering quality
        silhouette_avg = silhouette_score(
            tfidf_matrix.toarray(), best_labels, metric="euclidean"
        )

        # Calculate similarities using similarity threshold
        similarity_matrix = cosine_similarity(tfidf_matrix)
        similarity_matrix[similarity_matrix < similarity_threshold] = 0

        # Extract cluster information
        clusters = []
        for cluster_id in range(best_n_clusters):
            cluster_indices = np.where(best_labels == cluster_id)[0]

            # Skip if cluster is too small
            if len(cluster_indices) < min_cluster_size:
                continue

            # Calculate cluster terms
            cluster_tfidf = tfidf_matrix[cluster_indices].toarray()
            centroid = np.mean(cluster_tfidf, axis=0)

            # Get important terms with improved distinctiveness
            term_scores = []
            for idx, score in enumerate(centroid):
                if score > 0:
                    term = feature_names[idx]
                    cluster_freq = np.mean(cluster_tfidf[:, idx] > 0)
                    total_freq = np.mean(tfidf_matrix[:, idx].toarray() > 0)
                    distinctiveness = cluster_freq / (total_freq + 1e-10)

                    term_scores.append(
                        {
                            "term": term,
                            "score": float(score * distinctiveness),
                            "cluster_frequency": float(cluster_freq),
                            "total_frequency": float(total_freq),
                        }
                    )

            term_scores.sort(key=lambda x: x["score"], reverse=True)
            top_terms = term_scores[:20]

            # Get representative documents
            doc_similarities = []
            for idx in cluster_indices:
                doc_vector = tfidf_matrix[idx].toarray().flatten()
                sim_to_centroid = cosine_similarity(
                    doc_vector.reshape(1, -1), centroid.reshape(1, -1)
                )[0][0]

                doc_info = {
                    "title": display_data.iloc[idx]["Title"],
                    "date": display_data.iloc[idx]["date_of_report"],
                    "similarity": float(sim_to_centroid),
                    "summary": display_data.iloc[idx]["Content"][:500],
                }
                doc_similarities.append((idx, sim_to_centroid, doc_info))

            # Sort by similarity and get representative docs
            doc_similarities.sort(key=lambda x: x[1], reverse=True)
            representative_docs = [item[2] for item in doc_similarities]

            # Calculate cluster cohesion
            cluster_similarities = similarity_matrix[cluster_indices][
                :, cluster_indices
            ]
            cohesion = float(np.mean(cluster_similarities))

            clusters.append(
                {
                    "id": len(clusters),
                    "size": len(cluster_indices),
                    "cohesion": cohesion,
                    "terms": top_terms,
                    "documents": representative_docs,
                    "balance_ratio": max(
                        len(cluster_indices)
                        for cluster_indices in [
                            np.where(best_labels == i)[0]
                            for i in range(best_n_clusters)
                        ]
                    )
                    / min(
                        len(cluster_indices)
                        for cluster_indices in [
                            np.where(best_labels == i)[0]
                            for i in range(best_n_clusters)
                        ]
                    ),
                }
            )

        # Add cluster quality metrics to results
        metrics = {
            "silhouette_score": float(silhouette_avg),
            "calinski_score": float(
                calinski_harabasz_score(tfidf_matrix.toarray(), best_labels)
            ),
            "davies_score": float(
                davies_bouldin_score(tfidf_matrix.toarray(), best_labels)
            ),
            "balance_ratio": float(
                max(len(c["documents"]) for c in clusters)
                / min(len(c["documents"]) for c in clusters)
            ),
        }

        return {
            "n_clusters": len(clusters),
            "total_documents": len(processed_texts),
            "silhouette_score": float(silhouette_avg),
            "clusters": clusters,
            "vectorizer_type": vectorizer_type,
            "quality_metrics": metrics,
        }

    except Exception as e:
        logging.error(f"Error in semantic clustering: {e}", exc_info=True)
        raise


def create_document_identifier(row: pd.Series) -> str:
    """Create a unique identifier for a document based on its title and reference number"""
    title = str(row.get("Title", "")).strip()
    ref = str(row.get("ref", "")).strip()
    deceased = str(row.get("deceased_name", "")).strip()

    # Combine multiple fields to create a unique identifier
    identifier = f"{title}_{ref}_{deceased}"
    return identifier


def deduplicate_documents(data: pd.DataFrame) -> pd.DataFrame:
    """Deduplicate documents while preserving unique entries"""
    # Create unique identifiers
    data["doc_id"] = data.apply(create_document_identifier, axis=1)

    # Keep first occurrence of each unique document
    deduped_data = data.drop_duplicates(subset=["doc_id"])

    # Drop the temporary identifier column
    deduped_data = deduped_data.drop(columns=["doc_id"])

    return deduped_data


def format_date_uk(date_obj):
    """Convert datetime object to UK date format string"""
    if pd.isna(date_obj):
        return ""
    try:
        if isinstance(date_obj, str):
            # Try to parse string to datetime first
            date_obj = pd.to_datetime(date_obj)
        return date_obj.strftime("%d/%m/%Y")
    except:
        return str(date_obj)


def generate_extractive_summary(documents, max_length=500):
    """Generate extractive summary from cluster documents with traceability"""
    try:
        # Combine all document texts with source tracking
        all_sentences = []
        for doc in documents:
            sentences = sent_tokenize(doc["summary"])
            for sent in sentences:
                all_sentences.append(
                    {
                        "text": sent,
                        "source": doc["title"],
                        "date": format_date_uk(doc["date"]),  # Format date here
                    }
                )

        # Calculate sentence importance using TF-IDF
        vectorizer = TfidfVectorizer(stop_words="english")
        tfidf_matrix = vectorizer.fit_transform([s["text"] for s in all_sentences])

        # Calculate sentence scores
        sentence_scores = []
        for idx, sentence in enumerate(all_sentences):
            score = np.mean(tfidf_matrix[idx].toarray())
            sentence_scores.append((score, sentence))

        # Sort by importance and select top sentences
        sentence_scores.sort(reverse=True)
        summary_length = 0
        summary_sentences = []

        for score, sentence in sentence_scores:
            if summary_length + len(sentence["text"]) <= max_length:
                summary_sentences.append(
                    {
                        "text": sentence["text"],
                        "source": sentence["source"],
                        "date": sentence["date"],
                        "score": float(score),
                    }
                )
                summary_length += len(sentence["text"])
            else:
                break

        return summary_sentences

    except Exception as e:
        logging.error(f"Error in extractive summarization: {e}")
        return []


def generate_abstractive_summary(cluster_terms, documents, max_length=500):
    """Generate abstractive summary from cluster information with improved date handling"""
    try:
        # Extract key themes from terms
        top_themes = [term["term"] for term in cluster_terms[:5]]

        # Get document dates and format them with proper sorting
        dates = []
        for doc in documents:
            try:
                if doc["date"]:
                    date_obj = pd.to_datetime(doc["date"])
                    dates.append(date_obj)
            except:
                continue

        if dates:
            start_date = min(dates).strftime("%d/%m/%Y")
            end_date = max(dates).strftime("%d/%m/%Y")
            date_range = f"from {start_date} to {end_date}"
        else:
            date_range = ""

        # Extract key themes with better formatting
        main_themes = ", ".join(top_themes[:-1])
        if main_themes:
            themes_text = f"{main_themes} and {top_themes[-1]}"
        else:
            themes_text = top_themes[0] if top_themes else ""

        # Build better structured summary
        summary = f"This cluster contains {len(documents)} documents "
        if date_range:
            summary += f"{date_range} "
        summary += f"focused on {themes_text}. "

        # Add key patterns with improved statistics
        term_patterns = []
        for term in cluster_terms[5:8]:  # Get next 3 terms after main themes
            if term["cluster_frequency"] > 0:
                freq = term["cluster_frequency"] * 100
                # Add context based on frequency
                if freq > 75:
                    context = "very commonly"
                elif freq > 50:
                    context = "frequently"
                elif freq > 25:
                    context = "sometimes"
                else:
                    context = "occasionally"
                term_patterns.append(
                    f"{term['term']} ({context} appearing in {freq:.0f}% of documents)"
                )

        if term_patterns:
            summary += f"Common patterns include {', '.join(term_patterns)}. "

        # Add cluster distinctiveness if available
        if any(term["total_frequency"] < 0.5 for term in cluster_terms[:5]):
            distinctive_terms = [
                term["term"]
                for term in cluster_terms[:5]
                if term["total_frequency"] < 0.5
            ]
            if distinctive_terms:
                summary += f"This cluster is particularly distinctive in its discussion of {', '.join(distinctive_terms)}."

        # Truncate to max length while preserving complete sentences
        if len(summary) > max_length:
            summary = summary[:max_length]
            last_period = summary.rfind(".")
            if last_period > 0:
                summary = summary[: last_period + 1]

        return summary

    except Exception as e:
        logging.error(f"Error in abstractive summarization: {e}")
        return "Error generating summary"


def get_optimal_clustering_params(num_docs: int) -> Dict[str, int]:
    """Calculate optimal clustering parameters based on dataset size"""

    # Base parameters
    params = {
        "min_cluster_size": 2,  # Minimum starting point
        "max_features": 5000,  # Maximum vocabulary size
        "min_docs": 2,  # Minimum document frequency
        "max_docs": None,  # Maximum document frequency (will be calculated)
    }

    # Adjust minimum cluster size based on dataset size
    if num_docs < 10:
        params["min_cluster_size"] = 2
    elif num_docs < 20:
        params["min_cluster_size"] = 3
    elif num_docs < 50:
        params["min_cluster_size"] = 4
    else:
        params["min_cluster_size"] = 5

    # Adjust document frequency bounds
    params["min_docs"] = max(2, int(num_docs * 0.05))  # At least 5% of documents
    params["max_docs"] = min(
        int(num_docs * 0.95),  # No more than 95% of documents
        num_docs - params["min_cluster_size"],  # Leave room for at least one cluster
    )

    # Adjust feature count based on dataset size
    if num_docs < 20:
        params["max_features"] = 2000
    elif num_docs < 50:
        params["max_features"] = 3000
    elif num_docs < 100:
        params["max_features"] = 4000
    else:
        params["max_features"] = 5000

    return params


def display_cluster_analysis(cluster_results: Dict) -> None:
    """Display comprehensive cluster analysis results with quality metrics"""
    try:
        st.subheader("Document Clustering Analysis")

        # Overview metrics in two rows
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Number of Clusters", cluster_results["n_clusters"])
        with col2:
            st.metric("Total Documents", cluster_results["total_documents"])
        with col3:
            st.metric(
                "Average Cluster Size",
                round(
                    cluster_results["total_documents"] / cluster_results["n_clusters"],
                    1,
                ),
            )

        # Quality metrics
        st.subheader("Clustering Quality Metrics")
        metrics = cluster_results["quality_metrics"]

        qual_col1, qual_col2, qual_col3, qual_col4 = st.columns(4)

        with qual_col1:
            st.metric(
                "Silhouette Score",
                f"{metrics['silhouette_score']:.3f}",
                help="Measures how similar an object is to its own cluster compared to other clusters. Range: [-1, 1], higher is better.",
            )

        with qual_col2:
            st.metric(
                "Calinski-Harabasz Score",
                f"{metrics['calinski_score']:.0f}",
                help="Ratio of between-cluster to within-cluster dispersion. Higher is better.",
            )

        with qual_col3:
            st.metric(
                "Davies-Bouldin Score",
                f"{metrics['davies_score']:.3f}",
                help="Average similarity measure of each cluster with its most similar cluster. Lower is better.",
            )

        with qual_col4:
            st.metric(
                "Balance Ratio",
                f"{metrics['balance_ratio']:.1f}",
                help="Ratio of largest to smallest cluster size. Closer to 1 is better.",
            )

        # Display each cluster
        for cluster in cluster_results["clusters"]:
            with st.expander(
                f"Cluster {cluster['id']+1} ({cluster['size']} documents)",
                expanded=True,
            ):
                # Cluster metrics
                met_col1, met_col2 = st.columns(2)
                with met_col1:
                    st.metric(
                        "Cohesion Score",
                        f"{cluster['cohesion']:.3f}",
                        help="Average similarity between documents in the cluster",
                    )
                with met_col2:
                    st.metric(
                        "Size Percentage",
                        f"{(cluster['size'] / cluster_results['total_documents'] * 100):.1f}%",
                        help="Percentage of total documents in this cluster",
                    )

                # Terms analysis
                st.markdown("#### Key Terms")
                terms_df = pd.DataFrame(
                    [
                        {
                            "Term": term["term"],
                            "Frequency": f"{term['cluster_frequency']*100:.1f}%",
                            "Distinctiveness": f"{term['score']:.3f}",
                        }
                        for term in cluster["terms"][:10]
                    ]
                )
                st.dataframe(terms_df, hide_index=True)

                # Representative documents with formatted dates
                st.markdown("#### Representative Documents")
                for doc in cluster["documents"]:
                    st.markdown(
                        f"**{doc['title']}** (Similarity: {doc['similarity']:.2f})"
                    )
                    st.markdown(f"**Date**: {format_date_uk(doc['date'])}")
                    st.markdown(f"**Summary**: {doc['summary'][:300]}...")
                    st.markdown("---")

    except Exception as e:
        st.error(f"Error displaying cluster analysis: {str(e)}")
        logging.error(f"Display error: {str(e)}", exc_info=True)


def find_optimal_clusters(
    tfidf_matrix: sp.csr_matrix,
    min_clusters: int = 2,
    max_clusters: int = 10,
    min_cluster_size: int = 3,
) -> Tuple[int, np.ndarray]:
    """Find optimal number of clusters with relaxed constraints"""

    best_score = -1
    best_n_clusters = min_clusters
    best_labels = None

    # Store metrics for each clustering attempt
    metrics = []

    # Try different numbers of clusters
    for n_clusters in range(min_clusters, max_clusters + 1):
        try:
            # Perform clustering
            clustering = AgglomerativeClustering(
                n_clusters=n_clusters, metric="euclidean", linkage="ward"
            )

            labels = clustering.fit_predict(tfidf_matrix.toarray())

            # Calculate cluster sizes
            cluster_sizes = np.bincount(labels)

            # Skip if any cluster is too small
            if min(cluster_sizes) < min_cluster_size:
                continue

            # Calculate balance ratio (smaller is better)
            balance_ratio = max(cluster_sizes) / min(cluster_sizes)

            # Skip only if clusters are extremely imbalanced
            if balance_ratio > 10:  # Relaxed from 5 to 10
                continue

            # Calculate clustering metrics
            sil_score = silhouette_score(
                tfidf_matrix.toarray(), labels, metric="euclidean"
            )

            # Simplified scoring focused on silhouette and basic balance
            combined_score = sil_score * (
                1 - (balance_ratio / 20)
            )  # Relaxed balance penalty

            metrics.append(
                {
                    "n_clusters": n_clusters,
                    "silhouette": sil_score,
                    "balance_ratio": balance_ratio,
                    "combined_score": combined_score,
                    "labels": labels,
                }
            )

        except Exception as e:
            logging.warning(f"Error trying {n_clusters} clusters: {str(e)}")
            continue

    # If no configurations met the strict criteria, try to find the best available
    if not metrics:
        # Try again with minimal constraints
        for n_clusters in range(min_clusters, max_clusters + 1):
            try:
                clustering = AgglomerativeClustering(
                    n_clusters=n_clusters, metric="euclidean", linkage="ward"
                )

                labels = clustering.fit_predict(tfidf_matrix.toarray())
                sil_score = silhouette_score(
                    tfidf_matrix.toarray(), labels, metric="euclidean"
                )

                if sil_score > best_score:
                    best_score = sil_score
                    best_n_clusters = n_clusters
                    best_labels = labels

            except Exception as e:
                continue

        if best_labels is None:
            # If still no valid configuration, use minimum number of clusters
            clustering = AgglomerativeClustering(
                n_clusters=min_clusters, metric="euclidean", linkage="ward"
            )
            best_labels = clustering.fit_predict(tfidf_matrix.toarray())
            best_n_clusters = min_clusters
    else:
        # Use the best configuration from metrics
        best_metric = max(metrics, key=lambda x: x["combined_score"])
        best_n_clusters = best_metric["n_clusters"]
        best_labels = best_metric["labels"]

    return best_n_clusters, best_labels


def export_cluster_results(cluster_results: Dict) -> bytes:
    """Export cluster results with proper timestamp handling"""
    output = io.BytesIO()

    # Prepare export data with timestamp conversion
    export_data = {
        "metadata": {
            "total_documents": cluster_results["total_documents"],
            "number_of_clusters": cluster_results["n_clusters"],
            "silhouette_score": cluster_results["silhouette_score"],
        },
        "clusters": [],
    }

    # Convert cluster data
    for cluster in cluster_results["clusters"]:
        # Create a copy of cluster with converted documents
        cluster_export = cluster.copy()
        for doc in cluster_export["documents"]:
            # Ensure date is a string
            doc["date"] = str(doc["date"])

        export_data["clusters"].append(cluster_export)

    # Write JSON to BytesIO
    json.dump(export_data, io.TextIOWrapper(output, encoding="utf-8"), indent=2)
    output.seek(0)

    return output.getvalue()


def validate_data_state():
    """Check if valid data exists in session state"""
    return (
        "current_data" in st.session_state
        and st.session_state.current_data is not None
        and not st.session_state.current_data.empty
    )


def validate_model_state():
    """Check if valid topic model exists in session state"""
    return (
        "topic_model" in st.session_state and st.session_state.topic_model is not None
    )


def handle_no_data_state(section):
    """Handle state when no data is available"""
    st.warning("No data available. Please scrape reports or upload a file first.")
    uploaded_file = st.file_uploader(
        "Upload existing data file", type=["csv", "xlsx"], key=f"{section}_uploader"
    )

    if uploaded_file:
        try:
            df = (
                pd.read_csv(uploaded_file)
                if uploaded_file.name.endswith(".csv")
                else pd.read_excel(uploaded_file)
            )
            df = process_scraped_data(df)
            st.session_state.current_data = df
            st.rerun()
        except Exception as e:
            st.error(f"Error loading file: {str(e)}")


def handle_no_model_state():
    """Handle state when no topic model is available"""
    st.warning("Please run the clustering analysis first to view summaries.")
    if st.button("Go to Topic Modeling"):
        st.session_state.current_tab = "🔬 Topic Modeling"
        st.rerun()


def handle_error(error):
    """Handle application errors"""
    st.error("An error occurred")
    st.error(str(error))
    logging.error(f"Application error: {error}", exc_info=True)

    with st.expander("Error Details"):
        st.code(traceback.format_exc())

    st.warning("Recovery options:")
    st.markdown(
        """
    1. Clear data and restart
    2. Upload different data
    3. Check filter settings
    """
    )


def render_footer():
    """Render application footer"""
    st.markdown("---")
    st.markdown(
        """<div style='text-align: center'>
        <p>Built with Streamlit • Data from UK Judiciary</p>
        </div>""",
        unsafe_allow_html=True,
    )


def render_topic_modeling_tab(data: pd.DataFrame):
    """Render the topic modeling tab with enhanced visualization options"""
    st.header("Topic Modeling Analysis")

    if data is None or len(data) == 0:
        st.warning("No data available. Please scrape or upload data first.")
        return

    with st.sidebar:
        st.subheader("Vectorization Settings")
        vectorizer_type = st.selectbox(
            "Vectorization Method",
            ["tfidf", "bm25", "weighted"],
            help="Choose how to convert text to numerical features",
        )

        # BM25 specific parameters
        if vectorizer_type == "bm25":
            k1 = st.slider(
                "k1 parameter", 0.5, 3.0, 1.5, 0.1, help="Term saturation parameter"
            )
            b = st.slider(
                "b parameter",
                0.0,
                1.0,
                0.75,
                0.05,
                help="Length normalization parameter",
            )

        # Weighted TF-IDF parameters
        elif vectorizer_type == "weighted":
            tf_scheme = st.selectbox(
                "TF Weighting Scheme",
                ["raw", "log", "binary", "augmented"],
                help="How to weight term frequencies",
            )
            idf_scheme = st.selectbox(
                "IDF Weighting Scheme",
                ["smooth", "standard", "probabilistic"],
                help="How to weight inverse document frequencies",
            )

    # Analysis parameters
    st.subheader("Analysis Parameters")
    col1, col2 = st.columns(2)

    with col1:
        num_topics = st.slider(
            "Number of Topics",
            min_value=2,
            max_value=20,
            value=5,
            help="Number of distinct topics to identify",
        )

    with col2:
        max_features = st.slider(
            "Maximum Features",
            min_value=500,
            max_value=5000,
            value=1000,
            help="Maximum number of terms to consider",
        )

    # Get vectorizer parameters
    vectorizer_params = {}
    if vectorizer_type == "bm25":
        vectorizer_params.update({"k1": k1, "b": b})
    elif vectorizer_type == "weighted":
        vectorizer_params.update({"tf_scheme": tf_scheme, "idf_scheme": idf_scheme})

    if st.button("Run Analysis", type="primary"):
        with st.spinner("Performing topic analysis..."):
            try:
                # Create vectorizer
                vectorizer = get_vectorizer(
                    vectorizer_type=vectorizer_type,
                    max_features=max_features,
                    min_df=2,
                    max_df=0.95,
                    **vectorizer_params,
                )

                # Process text data
                docs = data["Content"].fillna("").apply(clean_text_for_modeling)

                # Create document-term matrix
                dtm = vectorizer.fit_transform(docs)
                feature_names = vectorizer.get_feature_names_out()

                # Fit LDA model
                lda = LatentDirichletAllocation(
                    n_components=num_topics, random_state=42, n_jobs=-1
                )

                doc_topics = lda.fit_transform(dtm)

                # Store model results
                st.session_state.topic_model = {
                    "lda": lda,
                    "vectorizer": vectorizer,
                    "feature_names": feature_names,
                    "doc_topics": doc_topics,
                }

                # Display results
                st.success("Topic analysis complete!")

                # Show topic words
                st.subheader("Topic Keywords")
                for idx, topic in enumerate(lda.components_):
                    top_words = [feature_names[i] for i in topic.argsort()[:-11:-1]]
                    st.markdown(f"**Topic {idx+1}:** {', '.join(top_words)}")

                # Display network visualization
                st.subheader("Topic Similarity Network")
                display_topic_network(lda, feature_names)

                # Show topic distribution
                st.subheader("Topic Distribution")
                topic_dist = doc_topics.mean(axis=0)
                topic_df = pd.DataFrame(
                    {
                        "Topic": [f"Topic {i+1}" for i in range(num_topics)],
                        "Proportion": topic_dist,
                    }
                )

                fig = px.bar(
                    topic_df,
                    x="Topic",
                    y="Proportion",
                    title="Topic Distribution Across Documents",
                )
                st.plotly_chart(fig, use_container_width=True)

                # Export options
                st.subheader("Export Results")
                if st.download_button(
                    "Download Topic Analysis Results",
                    data=export_topic_results(
                        lda, vectorizer, feature_names, doc_topics
                    ).encode(),
                    file_name="topic_analysis_results.json",
                    mime="application/json",
                ):
                    st.success("Results downloaded successfully!")

            except Exception as e:
                st.error(f"Error during analysis: {str(e)}")
                logging.error(f"Topic modeling error: {e}", exc_info=True)


def export_topic_results(lda_model, vectorizer, feature_names, doc_topics) -> str:
    """Export topic modeling results to JSON format"""
    results = {
        "topics": [],
        "model_params": {
            "n_topics": lda_model.n_components,
            "max_features": len(feature_names),
        },
        "topic_distribution": doc_topics.mean(axis=0).tolist(),
    }

    # Add topic details
    for idx, topic in enumerate(lda_model.components_):
        top_indices = topic.argsort()[:-11:-1]

        topic_words = [
            {"word": feature_names[i], "weight": float(topic[i])} for i in top_indices
        ]

        results["topics"].append(
            {"id": idx, "words": topic_words, "total_weight": float(topic.sum())}
        )

    return json.dumps(results, indent=2)


def render_summary_tab(cluster_results: Dict, original_data: pd.DataFrame) -> None:
    """Render cluster summaries and records with flexible column handling"""
    if not cluster_results or "clusters" not in cluster_results:
        st.warning("No cluster results available.")
        return

    st.write(
        f"Found {cluster_results['total_documents']} total documents in {cluster_results['n_clusters']} clusters"
    )

    for cluster in cluster_results["clusters"]:
        st.markdown(f"### Cluster {cluster['id']+1} ({cluster['size']} documents)")

        # Overview
        st.markdown("#### Overview")
        abstractive_summary = generate_abstractive_summary(
            cluster["terms"], cluster["documents"]
        )
        st.write(abstractive_summary)

        # Key terms table
        st.markdown("#### Key Terms")
        terms_df = pd.DataFrame(
            [
                {
                    "Term": term["term"],
                    "Frequency": f"{term['cluster_frequency']*100:.0f}%",
                }
                for term in cluster["terms"][:10]
            ]
        )
        st.dataframe(terms_df, hide_index=True)

        # Records
        st.markdown("#### Records")
        st.success(f"Showing {len(cluster['documents'])} matching documents")

        # Get the full records from original data
        doc_titles = [doc.get("title", "") for doc in cluster["documents"]]
        cluster_docs = original_data[original_data["Title"].isin(doc_titles)].copy()

        # Sort to match the original order
        title_to_position = {title: i for i, title in enumerate(doc_titles)}
        cluster_docs["sort_order"] = cluster_docs["Title"].map(title_to_position)
        cluster_docs = cluster_docs.sort_values("sort_order").drop("sort_order", axis=1)

        # Determine available columns
        available_columns = []
        column_config = {}

        # Always include URL and Title if available
        if "URL" in cluster_docs.columns:
            available_columns.append("URL")
            column_config["URL"] = st.column_config.LinkColumn("Report Link")

        if "Title" in cluster_docs.columns:
            available_columns.append("Title")
            column_config["Title"] = st.column_config.TextColumn("Title")

        # Add date if available
        if "date_of_report" in cluster_docs.columns:
            available_columns.append("date_of_report")
            column_config["date_of_report"] = st.column_config.DateColumn(
                "Date of Report", format="DD/MM/YYYY"
            )

        # Add optional columns if available
        optional_columns = [
            "ref",
            "deceased_name",
            "coroner_name",
            "coroner_area",
            "categories",
        ]
        for col in optional_columns:
            if col in cluster_docs.columns:
                available_columns.append(col)
                if col == "categories":
                    column_config[col] = st.column_config.ListColumn("Categories")
                else:
                    column_config[col] = st.column_config.TextColumn(
                        col.replace("_", " ").title()
                    )

        # Display the dataframe with available columns
        if available_columns:
            st.dataframe(
                cluster_docs[available_columns],
                column_config=column_config,
                hide_index=True,
                use_container_width=True,
            )
        else:
            st.warning("No displayable columns found in the data")

        st.markdown("---")


def main():
    """Updated main application entry point."""
    initialize_session_state()

    st.title("UK Judiciary PFD Reports Analysis")
    st.markdown(
        """
    This application analyzes Prevention of Future Deaths (PFD) reports from the UK Judiciary website.
    You can scrape new reports, analyze existing data, and explore thematic patterns.
    """
    )

    # Updated tab selection without topic modeling tab
    current_tab = st.radio(
        "Select section:",
        [
            "🔍 Scrape Reports",
            "📊 Analysis",
            "📝 Topic Analysis & Summaries",
            "🔬 BERT Analysis",  # New tab for BERT Analysis
        ],
        label_visibility="collapsed",
        horizontal=True,
        key="main_tab_selector",
    )

    st.markdown("---")

    try:
        if current_tab == "🔍 Scrape Reports":
            render_scraping_tab()

        elif current_tab == "📊 Analysis":
            if not validate_data_state():
                handle_no_data_state("analysis")
            else:
                render_analysis_tab(st.session_state.current_data)

        elif current_tab == "📝 Topic Analysis & Summaries":
            if not validate_data_state():
                handle_no_data_state("topic_summary")
            else:
                render_topic_summary_tab(st.session_state.current_data)

        elif current_tab == "🔬 BERT Analysis":
            if check_bert_password():
                render_bert_analysis_tab(st.session_state.current_data)

        # Sidebar data management
        with st.sidebar:
            st.header("Data Management")

            if hasattr(st.session_state, "data_source"):
                st.info(f"Current data: {st.session_state.data_source}")

            if st.button("Clear All Data"):
                for key in [
                    "current_data",
                    "scraped_data",
                    "uploaded_data",
                    "topic_model",
                    "data_source",
                ]:
                    if key in st.session_state:
                        del st.session_state[key]
                st.success("All data cleared")
                st.rerun()  # Use st.rerun() instead of experimental_rerun()

        render_footer()

    except Exception as e:
        handle_error(e)


def check_bert_password():
    """Returns `True` if the user had the correct password for BERT Analysis."""
    # Initialize session state variables if they don't exist
    if "bert_password_correct" not in st.session_state:
        st.session_state["bert_password_correct"] = False
    if "bert_password_attempted" not in st.session_state:
        st.session_state["bert_password_attempted"] = False

    # If already authenticated, return True
    if st.session_state["bert_password_correct"]:
        return True

    # Otherwise, show password input
    password = st.text_input(
        "Please enter the password for BERT Analysis",
        type="password",
        key="bert_password_input",
    )

    if st.button("Submit Password", key="bert_password_submit"):
        st.session_state["bert_password_attempted"] = True
        # Get the correct password from secrets
        correct_password = st.secrets.get("bert_password", "amazing246")

        if password == correct_password:
            st.session_state["bert_password_correct"] = True
            st.rerun()  # Use st.rerun() instead of experimental_rerun()
            return True
        else:
            st.error("Incorrect password. Please try again.")
            return False

    # Only show error if password has been attempted
    if (
        st.session_state["bert_password_attempted"]
        and not st.session_state["bert_password_correct"]
    ):
        st.error("Please enter the correct password to access BERT Analysis.")

    return False

    # Create PDF with matplotlib
    with PdfPages(output_filename) as pdf:
        # Title page
        plt.figure(figsize=(12, 8))
        plt.text(
            0.5, 0.5, "Theme Analysis Report", fontsize=24, ha="center", va="center"
        )
        plt.text(
            0.5,
            0.4,
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            fontsize=14,
            ha="center",
            va="center",
        )
        plt.axis("off")
        pdf.savefig()
        plt.close()

        # Summary statistics
        plt.figure(figsize=(12, 8))
        plt.text(0.5, 0.95, "Analysis Summary", fontsize=20, ha="center", va="center")

        # Document count
        doc_count = len(highlighted_texts)
        plt.text(0.1, 0.85, f"Total Documents Analyzed: {doc_count}", fontsize=14)

        # Theme count
        if not results_df.empty:
            theme_count = len(results_df)
            plt.text(0.1, 0.8, f"Total Theme Predictions: {theme_count}", fontsize=14)

            # Framework distribution
            plt.text(0.1, 0.7, "Framework Distribution:", fontsize=14)
            y_pos = 0.65
            for framework, count in results_df["Framework"].value_counts().items():
                plt.text(0.15, y_pos, f"{framework}: {count} themes", fontsize=12)
                y_pos -= 0.05

        plt.axis("off")
        pdf.savefig()
        plt.close()

        # Cannot directly include HTML in matplotlib, so just note that the highlights are available
        plt.figure(figsize=(12, 8))
        plt.text(
            0.5,
            0.5,
            "Highlighted texts are available in the web interface.\nThey cannot be directly included in this PDF.",
            fontsize=14,
            ha="center",
            va="center",
        )
        plt.axis("off")
        pdf.savefig()
        plt.close()

    return output_filename


def render_bert_analysis_tab(data: pd.DataFrame = None):
    """Modified render_bert_analysis_tab function with colored Theme column for text matching"""
    st.header("BERT-based Theme Analysis")

    # Check password before showing BERT analysis
    if check_bert_password():
        # Ensure the bert_results dictionary exists in session state
        if "bert_results" not in st.session_state:
            st.session_state.bert_results = {}

        # File upload section
        st.subheader("Upload Data")
        uploaded_file = st.file_uploader(
            "Upload CSV or Excel file for BERT Analysis",
            type=["csv", "xlsx"],
            help="Upload a file with reports for theme analysis",
            key="bert_file_uploader",
        )

        # If a file is uploaded, process it
        if uploaded_file is not None:
            try:
                if uploaded_file.name.endswith(".csv"):
                    uploaded_data = pd.read_csv(uploaded_file)
                else:
                    uploaded_data = pd.read_excel(uploaded_file)

                # Process the uploaded data
                uploaded_data = process_scraped_data(uploaded_data)

                # Update the data reference
                data = uploaded_data

                st.success("File uploaded and processed successfully!")
            except Exception as e:
                st.error(f"Error uploading file: {str(e)}")
                return

        # Check if data is available
        if data is None or len(data) == 0:
            st.warning(
                "No data available. Please upload a file or ensure existing data is loaded."
            )
            return

        # Column selection for analysis
        st.subheader("Select Analysis Column")

        # Find text columns (object/string type)
        text_columns = data.select_dtypes(include=["object"]).columns.tolist()

        # If no text columns found
        if not text_columns:
            st.error("No text columns found in the dataset.")
            return

        # Column selection with dropdown
        content_column = st.selectbox(
            "Choose the column to analyze:",
            options=text_columns,
            index=text_columns.index("Content") if "Content" in text_columns else 0,
            help="Select the column containing the text you want to analyze",
            key="bert_content_column",
        )

        # Filtering options
        st.subheader("Select Documents to Analyze")

        # Option to select all or specific records
        analysis_type = st.radio(
            "Analysis Type",
            ["All Reports", "Selected Reports"],
            horizontal=True,
            key="bert_analysis_type",
        )

        if analysis_type == "Selected Reports":
            # Multi-select for reports
            selected_indices = st.multiselect(
                "Choose specific reports to analyze",
                options=list(range(len(data))),
                format_func=lambda x: f"{data.iloc[x]['Title']} ({data.iloc[x]['date_of_report'].strftime('%d/%m/%Y') if pd.notna(data.iloc[x]['date_of_report']) else 'No date'})",
                key="bert_selected_indices",
            )
            selected_data = data.iloc[selected_indices] if selected_indices else None
        else:
            selected_data = data

        # Color scheme selection
        st.subheader("Theme Color Scheme")
        color_scheme = st.selectbox(
            "Choose Color Scheme",
            ["Pastel", "Vibrant", "Blues", "Earth Tones", "Distinct"],
            index=0,
            help="Select a color scheme for theme highlighting",
            key="theme_color_scheme",
        )

        # Define color schemes
        color_schemes = {
            "Pastel": [
                "#FFD580",
                "#FFECB3",
                "#E1F5FE",
                "#E8F5E9",
                "#F3E5F5",
                "#FFF3E0",
                "#E0F7FA",
                "#F1F8E9",
                "#FFF8E1",
                "#E8EAF6",
            ],
            "Vibrant": [
                "#FF5733",
                "#33FF57",
                "#3357FF",
                "#F333FF",
                "#FF33F3",
                "#33FFF3",
                "#FFFF33",
                "#FF3333",
                "#33FF33",
                "#3333FF",
            ],
            "Blues": [
                "#E3F2FD",
                "#BBDEFB",
                "#90CAF9",
                "#64B5F6",
                "#42A5F5",
                "#2196F3",
                "#1E88E5",
                "#1976D2",
                "#1565C0",
                "#0D47A1",
            ],
            "Earth Tones": [
                "#8D6E63",
                "#A1887F",
                "#BCAAA4",
                "#D7CCC8",
                "#EFEBE9",
                "#5D4037",
                "#4E342E",
                "#3E2723",
                "#BF360C",
                "#D84315",
            ],
            "Distinct": [
                "#E41A1C",
                "#377EB8",
                "#4DAF4A",
                "#984EA3",
                "#FF7F00",
                "#FFFF33",
                "#A65628",
                "#F781BF",
                "#999999",
                "#66C2A5",
            ],
        }

        # Analysis parameters
        st.subheader("Analysis Parameters")
        similarity_threshold = st.slider(
            "Similarity Threshold",
            min_value=0.3,
            max_value=0.9,
            value=0.65,
            step=0.05,
            help="Minimum similarity score for theme detection (higher = more strict)",
            key="bert_similarity_threshold",
        )

        # Analysis button
        run_analysis = st.button(
            "Run BERT Analysis", type="primary", key="bert_run_analysis"
        )

        # Run analysis if button is clicked
        if run_analysis:
            with st.spinner("Performing BERT Theme Analysis..."):
                try:
                    # Validate data selection
                    if selected_data is None or len(selected_data) == 0:
                        st.warning("No documents selected for analysis.")
                        return

                    # Initialize the theme analyzer
                    theme_analyzer = ThemeAnalyzer(
                        model_name="emilyalsentzer/Bio_ClinicalBERT"
                    )

                    # Set custom configuration
                    theme_analyzer.config[
                        "base_similarity_threshold"
                    ] = similarity_threshold

                    # Set color scheme
                    theme_analyzer.theme_colors = color_schemes[color_scheme]

                    # Perform analysis with highlighting
                    (
                        results_df,
                        highlighted_texts,
                    ) = theme_analyzer.create_detailed_results(
                        selected_data, content_column=content_column
                    )

                    # Generate timestamp for filenames
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    pdf_filename = f"theme_analysis_report_{timestamp}.pdf"

                    # Create PDF with comprehensive results
                    pdf_file = theme_analyzer.create_comprehensive_pdf(
                        results_df, highlighted_texts, pdf_filename
                    )

                    # Create HTML report
                    html_content = theme_analyzer._create_integrated_html_for_pdf(
                        results_df, highlighted_texts
                    )
                    html_filename = pdf_filename.replace(".pdf", ".html")

                    with open(html_filename, "w", encoding="utf-8") as f:
                        f.write(html_content)

                    # Create a mapping of theme keys to colors for consistent display
                    theme_colors = {}
                    for _, row in results_df.iterrows():
                        if "Framework" in row and "Theme" in row:
                            theme_key = f"{row['Framework']}_{row['Theme']}"
                            theme_colors[theme_key] = theme_analyzer._get_theme_color(
                                theme_key
                            )

                    # Save results to session state to ensure persistence
                    st.session_state.bert_results["results_df"] = results_df
                    st.session_state.bert_results[
                        "highlighted_texts"
                    ] = highlighted_texts
                    st.session_state.bert_results["pdf_filename"] = pdf_filename
                    st.session_state.bert_results["html_filename"] = html_filename
                    st.session_state.bert_results["theme_colors"] = theme_colors

                    st.success("Analysis complete!")

                except Exception as e:
                    st.error(f"Error during BERT analysis: {str(e)}")
                    logging.error(f"BERT analysis error: {e}", exc_info=True)

        # Always display results if they exist
        if (
            "bert_results" in st.session_state
            and st.session_state.bert_results.get("results_df") is not None
        ):
            results_df = st.session_state.bert_results["results_df"]
            highlighted_texts = st.session_state.bert_results["highlighted_texts"]
            pdf_filename = st.session_state.bert_results.get("pdf_filename")
            html_filename = st.session_state.bert_results.get("html_filename")
            theme_colors = st.session_state.bert_results.get("theme_colors", {})

            st.subheader("Analysis Summary")
            st.write(f"Total Records Analyzed: {len(highlighted_texts)}")
            st.write(f"Total Theme Predictions: {len(results_df)}")

            # Confidence distribution
            if "Confidence" in results_df.columns:
                st.write("\nConfidence Distribution:")
                st.write(results_df["Confidence"].value_counts())

            # Framework distribution
            st.write("\nFramework Distribution:")
            st.write(results_df["Framework"].value_counts())

            # Create columns for download buttons
            col1, col2, col3 = st.columns(3)

            # Generate consistent timestamp for filenames
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            with col1:
                # Excel download button
                excel_data = export_to_excel(results_df)
                st.download_button(
                    "📥 Download Excel Results",
                    data=excel_data,
                    file_name=f"bert_theme_analysis_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="bert_excel_download",
                )

            with col2:
                # PDF download button
                if pdf_filename and os.path.exists(pdf_filename):
                    with open(pdf_filename, "rb") as f:
                        pdf_data = f.read()

                    st.download_button(
                        "📄 Download PDF Report",
                        data=pdf_data,
                        file_name=os.path.basename(pdf_filename),
                        mime="application/pdf",
                        key="bert_pdf_download",
                    )
                else:
                    st.warning("PDF report not available")

            with col3:
                # HTML download button
                if html_filename and os.path.exists(html_filename):
                    with open(html_filename, "rb") as f:
                        html_data = f.read()

                    st.download_button(
                        "🌐 Download HTML Report",
                        data=html_data,
                        file_name=os.path.basename(html_filename),
                        mime="text/html",
                        key="bert_html_download",
                    )
                else:
                    st.warning("HTML report not available")

            # Show theme color legend
            if theme_colors:
                st.subheader("Theme Color Legend")
                legend_html = "<div style='display: flex; flex-wrap: wrap; gap: 10px; margin-bottom: 20px;'>"

                for theme_key, color in theme_colors.items():
                    theme_parts = theme_key.split("_", 1)
                    framework = theme_parts[0] if len(theme_parts) > 0 else ""
                    theme = theme_parts[1] if len(theme_parts) > 1 else theme_key

                    legend_html += f"""
                    <div style='display: flex; align-items: center; border: 1px solid #ddd; border-radius: 4px; padding: 5px; background-color: white; min-width: 300px;'>
                        <div style='width: 20px; height: 20px; background-color: {color}; border: 1px solid #666; margin-right: 10px;'></div>
                        <div>
                            <div style='font-weight: bold;'>{framework}</div>
                            <div>{theme}</div>
                        </div>
                    </div>
                    """

                legend_html += "</div>"
                st.markdown(legend_html, unsafe_allow_html=True)

            # Show results table with color indicators - CUSTOM HTML TABLE WITH COLORED THEME CELLS
            st.subheader("Identified Themes")

            # Create a styled HTML table with colored Theme column
            table_html = """
            <style>
                .theme-table {
                    width: 100%;
                    border-collapse: collapse;
                    font-family: Arial, sans-serif;
                }
                .theme-table th {
                    background-color: #4a86e8;
                    color: white;
                    font-weight: normal;
                    text-align: left;
                    padding: 8px 12px;
                    border: 1px solid #ddd;
                }
                .theme-table td {
                    padding: 8px 12px;
                    border: 1px solid #ddd;
                    vertical-align: middle;
                }
                .confidence-high {
                    background-color: #D5F5E3;
                }
                .confidence-medium {
                    background-color: #FCF3CF;
                }
                .confidence-low {
                    background-color: #FADBD8;
                }
            </style>
            
            <table class="theme-table">
                <tr>
                    <th>Framework</th>
                    <th>Theme</th>
                    <th>Confidence</th>
                    <th>Score</th>
                    <th>Matched Keywords</th>
                </tr>
            """

            # Add each row with appropriate styling
            for _, row in results_df.iterrows():
                framework = row["Framework"]
                theme = row["Theme"]
                confidence = row.get("Confidence", "")
                score = row.get("Combined Score", 0)
                keywords = row.get("Matched Keywords", "")

                # Get theme color
                theme_key = f"{framework}_{theme}"
                theme_color = theme_colors.get(theme_key, "#cccccc")

                # Set confidence class based on level
                confidence_class = ""
                if confidence == "High":
                    confidence_class = "confidence-high"
                elif confidence == "Medium":
                    confidence_class = "confidence-medium"
                elif confidence == "Low":
                    confidence_class = "confidence-low"

                # Add the row to the HTML table with colored Theme cell
                table_html += f"""
                <tr>
                    <td>{framework}</td>
                    <td style="background-color: {theme_color};">{theme}</td>
                    <td class="{confidence_class}">{confidence}</td>
                    <td>{score:.3f}</td>
                    <td>{keywords}</td>
                </tr>
                """

            table_html += "</table>"

            # Display the HTML table
            st.markdown(table_html, unsafe_allow_html=True)

            # Highlighted text preview
            st.subheader("Sample Highlighted Text")
            if highlighted_texts:
                # Create tabs for multiple documents
                doc_ids = list(highlighted_texts.keys())
                if len(doc_ids) > 5:
                    preview_ids = doc_ids[:5]  # Limit preview to 5 documents
                    st.info(
                        f"Showing 5 of {len(doc_ids)} documents. Download the full report for all results."
                    )
                else:
                    preview_ids = doc_ids

                tabs = st.tabs([f"Document {i+1}" for i in range(len(preview_ids))])

                for i, (tab, doc_id) in enumerate(zip(tabs, preview_ids)):
                    with tab:
                        # Get document title if available
                        doc_title = next(
                            (
                                row["Title"]
                                for _, row in results_df.iterrows()
                                if row.get("Record ID") == doc_id
                            ),
                            f"Document {i+1}",
                        )

                        st.markdown(f"### {doc_title}")

                        # Document themes with color indicators - match the main table style
                        doc_themes = results_df[results_df["Record ID"] == doc_id]
                        if not doc_themes.empty:
                            st.markdown("**Identified Themes:**")

                            # Create a matching table for document-specific themes
                            themes_table = """
                            <table class="theme-table">
                                <tr>
                                    <th>Framework</th>
                                    <th>Theme</th>
                                    <th>Confidence</th>
                                    <th>Matched Keywords</th>
                                </tr>
                            """

                            for _, theme_row in doc_themes.iterrows():
                                framework = theme_row["Framework"]
                                theme = theme_row["Theme"]
                                confidence = theme_row.get("Confidence", "N/A")
                                keywords = theme_row.get("Matched Keywords", "")

                                # Get theme color
                                theme_key = f"{framework}_{theme}"
                                theme_color = theme_colors.get(theme_key, "#cccccc")

                                # Set confidence class
                                confidence_class = ""
                                if confidence == "High":
                                    confidence_class = "confidence-high"
                                elif confidence == "Medium":
                                    confidence_class = "confidence-medium"
                                elif confidence == "Low":
                                    confidence_class = "confidence-low"

                                # Add row to document themes table
                                themes_table += f"""
                                <tr>
                                    <td>{framework}</td>
                                    <td style="background-color: {theme_color};">{theme}</td>
                                    <td class="{confidence_class}">{confidence}</td>
                                    <td>{keywords}</td>
                                </tr>
                                """

                            themes_table += "</table>"
                            st.markdown(themes_table, unsafe_allow_html=True)

                        # Highlighted text
                        st.markdown("**Highlighted Text:**")
                        st.markdown(highlighted_texts[doc_id], unsafe_allow_html=True)


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        st.error("Critical Error")
        st.error(str(e))
        logging.critical(f"Application crash: {e}", exc_info=True)
